from __future__ import annotations

import argparse
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path
import sys
import tempfile
from time import perf_counter
import tracemalloc

from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.core.database import Base  # noqa: E402
from app.models import (  # noqa: E402
    BackgroundJob,
    ImpactAnalysis,
    KnowledgeDocument,
    KnowledgeDocumentVersion,
    KnowledgeUnit,
    LineageEdge,
    LineageNode,
    MartField,
    MartTable,
    MartToYbtMapping,
    ProductScenario,
    Project,
    ScenarioBusinessMapping,
    ScenarioTechnicalLineage,
    SourceToMartMapping,
    TargetField,
    TargetTable,
)
from app.services.project_readiness import build_project_readiness  # noqa: E402
from app.services.task_queue import InlineTaskQueue  # noqa: E402


FULL_SCALE = {
    "tables": 5,
    "fields_per_table": 200,
    "scenarios": 20,
    "business_mappings": 2000,
    "technical_mappings": 1000,
    "double_layer_mappings": 1000,
    "knowledge_units": 5000,
    "lineage_edges": 5000,
    "impacts": 500,
}
SMALL_SCALE = {
    "tables": 2,
    "fields_per_table": 10,
    "scenarios": 3,
    "business_mappings": 30,
    "technical_mappings": 20,
    "double_layer_mappings": 20,
    "knowledge_units": 100,
    "lineage_edges": 100,
    "impacts": 10,
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Run a sanitized productization performance baseline")
    parser.add_argument("--small", action="store_true", help="Use the CI-sized dataset while exercising the same code paths")
    parser.add_argument("--output", type=Path, help="Write JSON metrics to this path")
    parser.add_argument("--excel-output", type=Path, help="Keep the generated formal Excel package")
    args = parser.parse_args()
    scale = SMALL_SCALE if args.small else FULL_SCALE
    metrics, workbook_bytes = run_baseline(scale, profile="small" if args.small else "full")
    rendered = json.dumps(metrics, ensure_ascii=False, indent=2, sort_keys=True)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered + "\n", encoding="utf-8")
    if args.excel_output:
        args.excel_output.parent.mkdir(parents=True, exist_ok=True)
        args.excel_output.write_bytes(workbook_bytes)
    print(rendered)


def run_baseline(scale: dict[str, int], *, profile: str) -> tuple[dict, bytes]:
    tracemalloc.start()
    with tempfile.TemporaryDirectory(prefix="ybt-performance-") as temporary:
        database_path = Path(temporary) / "baseline.db"
        engine = create_engine(f"sqlite:///{database_path}")
        Base.metadata.create_all(engine)
        try:
            with Session(engine) as db:
                started = perf_counter()
                ids = _create_dataset(db, scale)
                create_seconds = perf_counter() - started

                started = perf_counter()
                readiness = build_project_readiness(db, ids["project_id"])
                readiness_seconds = perf_counter() - started

                started = perf_counter()
                page = list(db.scalars(
                    select(TargetField)
                    .where(TargetField.project_id == ids["project_id"])
                    .order_by(TargetField.id)
                    .limit(100)
                ))
                list_seconds = perf_counter() - started

                started = perf_counter()
                search_hits = list(db.scalars(
                    select(KnowledgeUnit.id)
                    .where(
                        KnowledgeUnit.project_id == ids["project_id"],
                        KnowledgeUnit.normalized_content.contains("sanitized-regulatory"),
                    )
                    .limit(50)
                ))
                search_seconds = perf_counter() - started

                started = perf_counter()
                delivery_rows = _assemble_delivery_rows(db, ids["project_id"])
                assembly_seconds = perf_counter() - started

                started = perf_counter()
                rendered: dict[str, bytes] = {}

                def render_handler(handler_db: Session, _job: BackgroundJob) -> dict:
                    rendered["workbook"] = _render_formal_workbook(handler_db, ids["project_id"], delivery_rows)
                    return {"success_count": 1, "failed_count": 0, "byte_size": len(rendered["workbook"])}

                render_job = InlineTaskQueue().enqueue(
                    db,
                    job_type="performance_formal_excel_render",
                    institution_id=None,
                    project_id=ids["project_id"],
                    created_by=0,
                    idempotency_key=f"performance-baseline-{profile}",
                    payload_summary={"sanitized_fixture": True, "row_count": len(delivery_rows)},
                    handler=render_handler,
                )
                if render_job.status != "completed" or "workbook" not in rendered:
                    raise RuntimeError(f"Performance background render failed: {render_job.status}")
                workbook_bytes = rendered["workbook"]
                excel_seconds = perf_counter() - started

                background_job_count = db.scalar(
                    select(func.count(BackgroundJob.id)).where(BackgroundJob.project_id == ids["project_id"])
                ) or 0
                actual_counts = _actual_counts(db, ids["project_id"])
        finally:
            engine.dispose()
    _, peak_memory = tracemalloc.get_traced_memory()
    tracemalloc.stop()
    return {
        "profile": profile,
        "sanitized_fixture": True,
        "requested_scale": scale,
        "actual_counts": actual_counts,
        "metrics_seconds": {
            "create_test_data": round(create_seconds, 6),
            "project_readiness": round(readiness_seconds, 6),
            "field_list_page": round(list_seconds, 6),
            "knowledge_search": round(search_seconds, 6),
            "deliverable_assembly": round(assembly_seconds, 6),
            "excel_render": round(excel_seconds, 6),
        },
        "result_sizes": {
            "field_list_page": len(page),
            "knowledge_search_hits": len(search_hits),
            "deliverable_rows": len(delivery_rows),
            "excel_bytes": len(workbook_bytes),
        },
        "readiness": {
            "overall_status": readiness["overall_status"],
            "overall_score": readiness["score"],
            "dimension_count": len(readiness["dimensions"]),
        },
        "background_job_count": background_job_count,
        "background_render_status": render_job.status,
        "peak_memory_bytes": peak_memory,
        "workbook_sha256": sha256(workbook_bytes).hexdigest(),
    }, workbook_bytes


def _create_dataset(db: Session, scale: dict[str, int]) -> dict[str, int]:
    project = Project(name="公开脱敏性能基线", bank_name="示例银行", description="不包含真实机构或客户数据")
    db.add(project)
    db.flush()
    project_id = project.id

    table_rows = [
        {"id": index + 1, "project_id": project_id, "table_code": f"PERF_T{index + 1:02d}", "table_name": f"性能基线表 {index + 1}"}
        for index in range(scale["tables"])
    ]
    db.bulk_insert_mappings(TargetTable, table_rows)
    field_count = scale["tables"] * scale["fields_per_table"]
    field_rows = []
    for index in range(field_count):
        field_rows.append({
            "id": index + 1,
            "project_id": project_id,
            "target_table_id": index // scale["fields_per_table"] + 1,
            "field_code": f"PERF_FIELD_{index + 1:04d}",
            "field_name": f"性能字段 {index + 1}",
            "field_type": "VARCHAR(64)",
            "required_flag": index % 5 == 0,
            "field_definition": "公开脱敏性能基线字段定义",
            "regulatory_description": "仅用于容量与渲染测试",
        })
    db.bulk_insert_mappings(TargetField, field_rows)
    scenario_rows = [
        {"id": index + 1, "project_id": project_id, "scenario_code": f"SCENE_{index + 1:02d}", "scenario_name": f"示例场景 {index + 1}", "enabled": True, "sort_order": index}
        for index in range(scale["scenarios"])
    ]
    db.bulk_insert_mappings(ProductScenario, scenario_rows)

    combinations = [
        ((index % field_count) + 1, ((index // field_count) % scale["scenarios"]) + 1)
        for index in range(scale["business_mappings"])
    ]
    business_rows = [{
        "id": index + 1,
        "project_id": project_id,
        "target_field_id": field_id,
        "scenario_id": scenario_id,
        "business_definition": "公开脱敏业务定义",
        "final_content": f"sanitized business caliber {index + 1}",
        "business_confirm_status": "confirmed",
        "confidence_level": "confirmed",
    } for index, (field_id, scenario_id) in enumerate(combinations)]
    db.bulk_insert_mappings(ScenarioBusinessMapping, business_rows)
    technical_rows = [{
        "id": index + 1,
        "project_id": project_id,
        "target_field_id": combinations[index][0],
        "scenario_id": combinations[index][1],
        "business_mapping_id": index + 1,
        "source_system_name": "SAMPLE_SYSTEM",
        "source_schema_name": "SANITIZED",
        "source_table_english_name": f"PERF_SOURCE_{index % 50:02d}",
        "source_field_english_name": f"PERF_COLUMN_{index % 200:03d}",
        "processing_logic": "direct sanitized mapping",
        "final_content": f"sanitized technical lineage {index + 1}",
        "tech_confirm_status": "confirmed",
        "lineage_status": "verified",
        "confidence_level": "confirmed",
    } for index in range(scale["technical_mappings"])]
    db.bulk_insert_mappings(ScenarioTechnicalLineage, technical_rows)

    mart = MartTable(id=1, project_id=project_id, table_code="PERF_MART", table_name="公开脱敏性能集市")
    db.add(mart)
    db.flush()
    double_count = scale["double_layer_mappings"]
    mart_field_count = max(1, (double_count + 1) // 2)
    db.bulk_insert_mappings(MartField, [{
        "id": index + 1,
        "project_id": project_id,
        "mart_table_id": mart.id,
        "field_code": f"MART_FIELD_{index + 1:04d}",
        "field_name": f"集市字段 {index + 1}",
        "field_type": "VARCHAR(64)",
        "is_existing": True,
    } for index in range(mart_field_count)])
    source_count = double_count // 2
    ybt_count = double_count - source_count
    db.bulk_insert_mappings(SourceToMartMapping, [{
        "id": index + 1,
        "project_id": project_id,
        "mart_field_id": index % mart_field_count + 1,
        "mapping_name": f"源到集市 {index + 1}",
        "mapping_status": "approved",
        "final_content": "sanitized source-to-mart mapping",
        "lineage_status": "verified",
    } for index in range(source_count)])
    db.bulk_insert_mappings(MartToYbtMapping, [{
        "id": index + 1,
        "project_id": project_id,
        "target_field_id": index % field_count + 1,
        "mart_field_id": index % mart_field_count + 1,
        "mapping_name": f"集市到一表通 {index + 1}",
        "mapping_status": "approved",
        "final_content": "sanitized mart-to-ybt mapping",
        "lineage_status": "verified",
    } for index in range(ybt_count)])

    document = KnowledgeDocument(
        id=1,
        project_id=project_id,
        file_name="sanitized-performance-knowledge.txt",
        file_type="txt",
        source_type="generated_fixture",
        storage_path="memory://sanitized-performance-knowledge",
        knowledge_type="regulatory_policy",
        knowledge_scope="project",
        document_status="active",
    )
    version = KnowledgeDocumentVersion(
        id=1,
        project_id=project_id,
        document_id=1,
        version_no=1,
        file_name=document.file_name,
        storage_path=document.storage_path,
        file_hash="a" * 64,
        parse_status="completed",
    )
    db.add_all([document, version])
    db.flush()
    db.bulk_insert_mappings(KnowledgeUnit, [{
        "id": index + 1,
        "project_id": project_id,
        "document_id": 1,
        "document_version_id": 1,
        "knowledge_type": "regulatory_policy",
        "knowledge_scope": "project",
        "unit_type": "paragraph",
        "title": f"公开知识单元 {index + 1}",
        "content": f"sanitized-regulatory performance knowledge unit {index + 1}",
        "normalized_content": f"sanitized-regulatory performance knowledge unit {index + 1}",
        "source_file_name": document.file_name,
        "target_field_code": f"PERF_FIELD_{index % field_count + 1:04d}",
        "scenario_id": index % scale["scenarios"] + 1,
        "tags_json": ["sanitized", "performance"],
        "metadata_json": {"fixture": True},
        "confidentiality_level": "public",
        "enabled": True,
        "content_hash": sha256(f"knowledge-{index}".encode()).hexdigest(),
    } for index in range(scale["knowledge_units"])])

    node_count = max(2, min(scale["lineage_edges"] + 1, field_count * 2))
    db.bulk_insert_mappings(LineageNode, [{
        "id": index + 1,
        "project_id": project_id,
        "node_type": "column",
        "logical_name": f"SANITIZED.PERF_COLUMN_{index + 1}",
        "schema_name": "SANITIZED",
        "table_name": f"PERF_TABLE_{index % scale['tables'] + 1}",
        "column_name": f"PERF_COLUMN_{index + 1}",
        "temporary_flag": False,
        "unresolved_flag": False,
        "metadata_json": {"fixture": True},
    } for index in range(node_count)])
    db.bulk_insert_mappings(LineageEdge, [{
        "id": index + 1,
        "project_id": project_id,
        "script_file_version_id": 1,
        "source_node_id": index % node_count + 1,
        "target_node_id": (index + 1) % node_count + 1,
        "edge_type": "derives_from",
        "transformation_type": "direct",
        "transformation_expression": "sanitized direct mapping",
        "confidence_level": "high",
        "evidence_json": {"fixture": True},
        "enabled": True,
    } for index in range(scale["lineage_edges"])])
    db.bulk_insert_mappings(ImpactAnalysis, [{
        "id": index + 1,
        "project_id": project_id,
        "change_set_id": index + 1,
        "status": "completed",
        "severity": ("low", "medium", "high", "critical")[index % 4],
        "affected_target_field_ids_json": [index % field_count + 1],
        "affected_lineage_edge_ids_json": [index % scale["lineage_edges"] + 1],
        "summary_json": {"fixture": True, "index": index + 1},
        "open_questions_json": [],
    } for index in range(scale["impacts"])])
    db.commit()
    return {"project_id": project_id}


def _assemble_delivery_rows(db: Session, project_id: int) -> list[dict]:
    fields = list(db.scalars(
        select(TargetField).where(TargetField.project_id == project_id).order_by(TargetField.id)
    ))
    business_counts = dict(db.execute(
        select(ScenarioBusinessMapping.target_field_id, func.count(ScenarioBusinessMapping.id))
        .where(ScenarioBusinessMapping.project_id == project_id)
        .group_by(ScenarioBusinessMapping.target_field_id)
    ).all())
    technical_counts = dict(db.execute(
        select(ScenarioTechnicalLineage.target_field_id, func.count(ScenarioTechnicalLineage.id))
        .where(ScenarioTechnicalLineage.project_id == project_id)
        .group_by(ScenarioTechnicalLineage.target_field_id)
    ).all())
    return [{
        "field_code": field.field_code,
        "field_name": field.field_name,
        "field_type": field.field_type,
        "definition": field.field_definition,
        "business_mapping_count": business_counts.get(field.id, 0),
        "technical_mapping_count": technical_counts.get(field.id, 0),
    } for field in fields]


def _render_formal_workbook(db: Session, project_id: int, delivery_rows: list[dict]) -> bytes:
    workbook = Workbook(write_only=True)
    overview = workbook.create_sheet("交付概览")
    overview.append(["项目", "公开脱敏性能基线"])
    overview.append(["目标字段数", len(delivery_rows)])
    fields = workbook.create_sheet("业务口径及技术溯源表")
    fields.append(["字段代码", "字段名称", "类型", "定义", "业务口径数", "技术溯源数"])
    for item in delivery_rows:
        fields.append(list(item.values()))
    knowledge = workbook.create_sheet("知识单元")
    knowledge.append(["标题", "内容", "来源文件"])
    for item in db.execute(
        select(KnowledgeUnit.title, KnowledgeUnit.content, KnowledgeUnit.source_file_name)
        .where(KnowledgeUnit.project_id == project_id)
        .order_by(KnowledgeUnit.id)
    ):
        knowledge.append(list(item))
    lineage = workbook.create_sheet("字段级血缘")
    lineage.append(["源节点", "目标节点", "类型", "转换"])
    for item in db.execute(
        select(LineageEdge.source_node_id, LineageEdge.target_node_id, LineageEdge.edge_type, LineageEdge.transformation_expression)
        .where(LineageEdge.project_id == project_id)
        .order_by(LineageEdge.id)
    ):
        lineage.append(list(item))
    impacts = workbook.create_sheet("脚本变更影响")
    impacts.append(["影响编号", "严重程度", "状态", "摘要"])
    for item in db.scalars(
        select(ImpactAnalysis).where(ImpactAnalysis.project_id == project_id).order_by(ImpactAnalysis.id)
    ):
        impacts.append([item.id, item.severity, item.status, json.dumps(item.summary_json, ensure_ascii=False)])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _actual_counts(db: Session, project_id: int) -> dict[str, int]:
    models = {
        "target_tables": TargetTable,
        "target_fields": TargetField,
        "scenarios": ProductScenario,
        "business_mappings": ScenarioBusinessMapping,
        "technical_mappings": ScenarioTechnicalLineage,
        "source_to_mart_mappings": SourceToMartMapping,
        "mart_to_ybt_mappings": MartToYbtMapping,
        "knowledge_units": KnowledgeUnit,
        "lineage_edges": LineageEdge,
        "impacts": ImpactAnalysis,
    }
    counts = {
        key: db.scalar(select(func.count(model.id)).where(model.project_id == project_id)) or 0
        for key, model in models.items()
    }
    counts["double_layer_mappings"] = counts["source_to_mart_mappings"] + counts["mart_to_ybt_mappings"]
    return counts


if __name__ == "__main__":
    main()
