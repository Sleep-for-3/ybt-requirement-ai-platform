from collections.abc import Iterator
from contextlib import contextmanager
from io import BytesIO
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.main import app
from app.services.storage import get_storage_service
from app.services.deliverables.workbook import render_workbook


def test_template_version_render_history_reuse_and_delivery_lifecycle(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_DIR", str(tmp_path / "storage")); get_storage_service.cache_clear()
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "正式交付项目"})
        table = _post(client, "/api/target-tables", {"project_id": project["id"], "table_code": "YBT_CUSTOMER", "table_name": "客户信息"})
        field = _post(client, "/api/fields", {"project_id": project["id"], "target_table_id": table["id"], "field_code": "CERT_TYPE", "field_name": "客户证件类型", "field_type": "VARCHAR", "regulatory_description": "客户证件类型监管定义"})
        scenario = _post(client, f"/api/projects/{project['id']}/scenarios", {"scenario_code": "DEBIT", "scenario_name": "借记卡"})
        business = _post(client, f"/api/target-fields/{field['id']}/scenarios/{scenario['id']}/business-mapping", {"final_content": "人工确认前的业务最终内容", "business_definition": "借记卡客户主证件类型"})
        technical = _post(client, f"/api/target-fields/{field['id']}/scenarios/{scenario['id']}/technical-lineage", {"business_mapping_id": business["id"], "source_system_name": "ECIF", "source_schema_name": "ODS", "source_table_english_name": "ECIF_CUSTOMER", "source_field_english_name": "CERT_TYPE", "final_content": "ODS.ECIF_CUSTOMER.CERT_TYPE"})
        _post(client, f"/api/mappings/scenario_technical/{technical['id']}/evidence", {"evidence_type": "manual_note", "source_name": "restricted 技术访谈", "quoted_content": "restricted 原文包含账号 6222000012345678", "evidence_summary": "来源字段已脱敏核验"})

        template_bytes = _template_workbook()
        upload = client.post(f"/api/projects/{project['id']}/deliverable-templates/upload", data={"template_name": "银行正式交付模板", "template_type": "full_delivery_package"}, files={"file": ("正式模板.xlsx", template_bytes, XLSX)})
        assert upload.status_code == 201, upload.text
        uploaded = upload.json(); version_id = uploaded["version"]["id"]; template_id = uploaded["template"]["id"]
        assert uploaded["version"]["sheet_config_json"][0]["merged_cells"] == ["A1:C1"]
        repeated = client.post(f"/api/projects/{project['id']}/deliverable-templates/upload", data={"template_id": template_id, "template_name": "银行正式交付模板", "template_type": "full_delivery_package"}, files={"file": ("正式模板.xlsx", template_bytes, XLSX)})
        assert repeated.status_code == 201; assert repeated.json()["version"]["id"] == version_id

        configured = _post(client, f"/api/deliverable-template-versions/{version_id}/configure", {"sheet_mappings": [{"business_section": "target_field", "sheet_name": "业务口径及技术溯源表", "header_row_start": 2, "header_row_end": 2, "data_start_row": 3, "columns": [{"business_field": "target_field_code", "excel_column": "A", "required": True}, {"business_field": "target_field_name", "excel_column": "B", "required": True}]}, {"business_section": "evidence", "sheet_name": "证据引用清单", "data_start_row": 2, "columns": [{"business_field": "evidence_type", "excel_column": "A"}, {"business_field": "evidence_summary", "excel_column": "B"}]}]})
        assert len(configured["column_mappings"]) == 4
        preview = client.post(f"/api/deliverable-template-versions/{version_id}/preview-render")
        assert preview.status_code == 200
        preview_book = load_workbook(BytesIO(preview.content), data_only=False); preview_sheet = preview_book["业务口径及技术溯源表"]
        assert preview_sheet["A3"].value == "FIELD_001"; assert preview_sheet["A2"].font.bold; assert preview_sheet["D3"].value == "=1+1"; assert "A1:C1" in [str(item) for item in preview_sheet.merged_cells.ranges]; assert preview_sheet.freeze_panes == "A3"; assert len(preview_sheet.data_validations.dataValidation) == 1

        history = client.post(f"/api/projects/{project['id']}/historical-calibers/upload", data={"document_type": "business_traceability", "import_name": "脱敏历史业务口径"}, files={"file": ("历史业务口径.xlsx", _history_workbook(), XLSX)})
        assert history.status_code == 201, history.text; history_item = history.json()["items"][0]; assert history_item["match_status"] == "matched"
        reused = _post(client, f"/api/historical-caliber-items/{history_item['id']}/reuse", {})
        assert reused["final_content_overwritten"] is False
        current = client.get(f"/api/scenario-business-mappings/{business['id']}").json()
        assert current["final_content"] == "人工确认前的业务最终内容"; assert "历史建议" in current["ai_generated_content"]

        question = _post(client, f"/api/projects/{project['id']}/questions", {"target_table_id": table["id"], "target_field_id": field["id"], "scenario_id": scenario["id"], "question_type": "source_field", "question_text": "来源字段需科技确认", "priority": "high", "assigned_role": "technical_analyst"})
        other_project = _post(client, "/api/projects", {"name": "隔离项目"}); other_table = _post(client, "/api/target-tables", {"project_id": other_project["id"], "table_code": "OTHER", "table_name": "其他表"}); other_field = _post(client, "/api/fields", {"project_id": other_project["id"], "target_table_id": other_table["id"], "field_code": "OTHER_FIELD", "field_name": "其他字段"})
        cross_reference = client.post(f"/api/projects/{project['id']}/questions", json={"target_table_id": table["id"], "target_field_id": other_field["id"], "question_text": "不应允许的跨项目引用"})
        assert cross_reference.status_code == 404
        cross_comparison = client.post(f"/api/projects/{project['id']}/caliber-comparisons", json={"target_field_id": other_field["id"], "left": {}, "right": {}})
        assert cross_comparison.status_code == 404
        readiness = client.get(f"/api/target-fields/{field['id']}/delivery-readiness").json(); assert readiness["status"] == "blocked"; assert readiness["open_question_count"] == 1
        package = _post(client, f"/api/projects/{project['id']}/deliverables", {"package_name": "客户信息正式交付包", "target_table_id": table["id"], "template_version_id": version_id})
        generated = _post(client, f"/api/deliverables/{package['id']}/generate", {}); assert generated["job"]["status"] == "completed"
        validation = _post(client, f"/api/deliverables/{package['id']}/validate", {}); assert validation["error_count"] > 0; assert any(item["code"] == "high_priority_question" for item in validation["issues"])
        assert client.post(f"/api/deliverables/{package['id']}/approve").status_code == 409
        _post(client, f"/api/questions/{question['id']}/answer", {"resolution_text": "已确认使用 CERT_TYPE"}); _post(client, f"/api/questions/{question['id']}/accept", {})
        rendered = _post(client, f"/api/deliverables/{package['id']}/render", {}); assert rendered["file_id"]
        downloaded = client.get(f"/api/deliverables/{package['id']}/download"); assert downloaded.status_code == 200; rendered_book = load_workbook(BytesIO(downloaded.content)); assert rendered_book["业务口径及技术溯源表"]["A3"].value == "CERT_TYPE"; assert rendered_book["证据引用清单"]["B2"].value == "来源字段已脱敏核验"; assert "6222000012345678" not in " ".join(str(cell.value or "") for sheet in rendered_book.worksheets for row in sheet.iter_rows() for cell in row)


def test_semantic_comparison_ignores_format_only_changes() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "版本比较项目"})
        comparison = _post(client, f"/api/projects/{project['id']}/caliber-comparisons", {"left": {"business_content": "仅包含 当前有效客户。"}, "right": {"business_content": "仅包含当前有效客户"}})
        assert comparison["result_json"]["business_content"]["difference_type"] == "unchanged"


def test_template_renderer_expands_scenarios_horizontally_and_sources_vertically() -> None:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "场景"; sheet["A2"] = "场景"; source = workbook.create_sheet("来源"); source["A2"] = "来源"; stream = BytesIO(); workbook.save(stream)
    sheets = [SimpleNamespace(id=1, enabled=True, sheet_name="场景", business_section="scenario_business_mapping", data_start_row=2, repeat_direction="horizontal"), SimpleNamespace(id=2, enabled=True, sheet_name="来源", business_section="source_to_mart", data_start_row=2, repeat_direction="vertical")]
    columns = [SimpleNamespace(template_sheet_mapping_id=1, business_field="scenario_name", excel_column="A", default_value=None, required=True, write_mode="repeat_by_scenario", merge_strategy="none"), SimpleNamespace(template_sheet_mapping_id=2, business_field="source_field_name", excel_column="A", default_value=None, required=True, write_mode="repeat_by_source", merge_strategy="none")]
    rendered, warnings = render_workbook(stream.getvalue(), sheets, columns, {"scenario_business_mapping": [{"scenario_name": "借记卡"}, {"scenario_name": "信用卡"}], "source_to_mart": [{"source_field_name": "cert_type"}, {"source_field_name": "id_type"}]})
    result = load_workbook(BytesIO(rendered)); assert result["场景"]["A2"].value == "借记卡"; assert result["场景"]["B2"].value == "信用卡"; assert result["来源"]["A2"].value == "cert_type"; assert result["来源"]["A3"].value == "id_type"; assert warnings == []


XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _template_workbook() -> bytes:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "业务口径及技术溯源表"; sheet.merge_cells("A1:C1"); sheet["A1"] = "银行正式交付模板"; sheet["A2"] = "字段代码"; sheet["B2"] = "字段名称"; sheet["C2"] = "可信等级"; sheet["D3"] = "=1+1"; sheet.freeze_panes = "A3"; sheet["A2"].font = Font(bold=True); sheet["A2"].fill = PatternFill("solid", fgColor="D9EAF7"); sheet.column_dimensions["B"].width = 28; sheet.row_dimensions[2].height = 25
    validation = DataValidation(type="list", formula1='"confirmed,inferred"'); validation.add("C3:C100"); sheet.add_data_validation(validation)
    evidence = workbook.create_sheet("证据引用清单"); evidence.append(["证据类型", "脱敏摘要"])
    stream = BytesIO(); workbook.save(stream); return stream.getvalue()


def _history_workbook() -> bytes:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "历史业务口径"; sheet.append(["字段代码", "字段名称", "业务场景", "业务口径", "来源系统", "来源schema", "来源表英文名", "来源字段英文名"]); sheet.append(["CERT_TYPE", "客户证件类型", "借记卡", "历史业务口径建议", "ECIF", "ODS", "ECIF_CUSTOMER", "CERT_TYPE"]); stream = BytesIO(); workbook.save(stream); return stream.getvalue()


@contextmanager
def _client() -> Iterator[TestClient]:
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool); Base.metadata.create_all(engine); factory = sessionmaker(bind=engine, autoflush=False)
    def override() -> Iterator[Session]:
        session = factory()
        try: yield session
        finally: session.close()
    app.dependency_overrides[get_db] = override
    try:
        with TestClient(app) as client: yield client
    finally:
        app.dependency_overrides.clear(); Base.metadata.drop_all(engine); get_storage_service.cache_clear()


def _post(client, path, payload):
    response = client.post(path, json=payload); assert response.status_code in {200, 201}, response.text; return response.json()
