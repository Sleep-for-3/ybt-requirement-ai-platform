from io import BytesIO
from pathlib import Path
from collections.abc import Iterator
from contextlib import contextmanager

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.main import app
from app.services.template_parser.traceability_excel_parser import TraceabilityExcelParser


def test_parser_recognizes_merged_multilevel_headers_and_partial_scenarios(tmp_path: Path) -> None:
    path = tmp_path / "脱敏业务口径及溯源表.xlsx"
    path.write_bytes(_workbook_bytes())

    output = TraceabilityExcelParser().parse(str(path))

    assert output.sheet_count == 1
    result = output.results[0]
    assert result.header_start_row == 2
    assert result.header_end_row == 3
    assert {group["scenario_name"] for group in result.scenario_groups} == {"借记卡", "信用卡", "贷款产品"}
    row = result.parsed_rows[0]
    assert row["fixed"]["field_code"] == "CARD_PRODUCT_ID"
    assert row["scenarios"]["借记卡"]["business"]["business_definition"] == "借记卡产品唯一编号"
    assert row["scenarios"]["借记卡"]["technical"]["source_system_name"] == "借记卡系统"
    assert row["scenarios"]["信用卡"]["business"]["business_definition"] == "信用卡产品编号"
    assert row["scenarios"]["信用卡"]["technical"] == {}
    assert row["scenarios"]["贷款产品"]["business"] == {}
    assert row["scenarios"]["贷款产品"]["technical"]["source_table_english_name"] == "LOAN_PRODUCT"
    assert row["sources"]["fixed.field_code"] == "业务口径!A4"


def test_upload_preview_and_apply_create_scenarios_mappings_lineage_and_knowledge() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "历史口径导入项目"})
        response = client.post(
            "/api/traceability-templates/upload",
            data={"project_id": str(project["id"])},
            files={
                "file": (
                    "脱敏业务口径及溯源表.xlsx",
                    _workbook_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        response.raise_for_status()
        uploaded = response.json()
        assert uploaded["parse_status"] == "success"
        assert {item["scenario_name"] for item in uploaded["detected_scenarios"]} == {"借记卡", "信用卡", "贷款产品"}

        preview = _get(client, f"/api/traceability-templates/{uploaded['template_id']}/preview")
        assert preview["results"][0]["header_start_row"] == 2
        assert preview["results"][0]["parsed_rows_json"][0]["fixed"]["field_code"] == "CARD_PRODUCT_ID"

        applied = _post(client, f"/api/traceability-templates/{uploaded['template_id']}/apply", {})
        assert applied["created_fields"] == 1
        assert applied["created_scenarios"] == 3
        assert applied["created_business_mappings"] == 2
        assert applied["created_technical_lineages"] == 2
        assert applied["created_knowledge_items"] >= 5

        reapplied = _post(client, f"/api/traceability-templates/{uploaded['template_id']}/apply", {})
        assert reapplied["created_fields"] == 0
        assert reapplied["created_scenarios"] == 0
        assert reapplied["created_business_mappings"] == 0
        assert reapplied["created_technical_lineages"] == 0
        assert reapplied["created_knowledge_items"] == 0

        scenarios = _get(client, f"/api/projects/{project['id']}/scenarios")
        fields = _get(client, f"/api/fields?project_id={project['id']}")
        field_id = fields[0]["id"]
        businesses = _get(client, f"/api/target-fields/{field_id}/scenario-business-mappings")
        lineages = _get(client, f"/api/target-fields/{field_id}/scenario-technical-lineages")
        knowledge = _get(client, f"/api/projects/{project['id']}/knowledge/items?target_field_code=CARD_PRODUCT_ID")
        assert len(scenarios) == 3
        assert len(businesses) == 2
        assert len(lineages) == 2
        assert {item["business_definition"] for item in businesses} == {"借记卡产品唯一编号", "信用卡产品编号"}
        assert {item["source_system_name"] for item in lineages} == {"借记卡系统", "信贷系统"}
        assert all(item["source_cell_range"] for item in knowledge)


def test_export_traceability_workbook_has_fixed_and_merged_scenario_headers() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "溯源表导出项目"})
        upload = client.post(
            "/api/traceability-templates/upload",
            data={"project_id": str(project["id"])},
            files={"file": ("脱敏历史口径.xlsx", _workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        upload.raise_for_status()
        _post(client, f"/api/traceability-templates/{upload.json()['template_id']}/apply", {})

        field = _get(client, f"/api/fields?project_id={project['id']}")[0]
        lineages = _get(client, f"/api/target-fields/{field['id']}/scenario-technical-lineages")
        debit_lineage = next(item for item in lineages if item["source_system_name"] == "借记卡系统")
        _put(
            client,
            f"/api/scenario-technical-lineages/{debit_lineage['id']}",
            {"final_content": "人工最终技术口径：借记卡产品字段直接取值。"},
        )

        response = client.get(f"/api/projects/{project['id']}/export/traceability-workbook")
        response.raise_for_status()
        assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook["业务口径及技术溯源"]
        first_headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        second_headers = [sheet.cell(2, column).value for column in range(1, sheet.max_column + 1)]
        assert first_headers[0] == "数据项编码"
        assert "业务口径-借记卡" in first_headers
        assert "溯源-借记卡" in first_headers
        assert "来源系统" in second_headers
        assert "来源表英文名" in second_headers
        assert "来源字段英文名" in second_headers
        assert "处理逻辑" in second_headers
        assert "技术口径确认人" in second_headers
        assert sheet.freeze_panes == "L3"
        assert "人工最终技术口径：借记卡产品字段直接取值。" in [
            sheet.cell(3, column).value for column in range(1, sheet.max_column + 1)
        ]
        assert any(cell_range.min_row == 1 and cell_range.max_row == 1 for cell_range in sheet.merged_cells.ranges)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "业务口径"
    sheet["A1"] = "脱敏模拟业务口径及技术溯源表"

    fixed_headers = [
        "数据项编码", "数据项名称", "数据类别", "数据格式", "字段业务定义（监管原始口径）",
        "字段业务定义（监管定义细化）", "报表名称", "字段名称", "EAST口径", "字段业务定义（行内）", "备注",
    ]
    for column, value in enumerate(fixed_headers, start=1):
        sheet.cell(2, column, value)
        sheet.merge_cells(start_row=2, start_column=column, end_row=3, end_column=column)

    business_headers = ["字段业务定义", "源系统截图", "源系统改造", "外部数据", "手工补录", "业务口径确认人", "备注"]
    technical_headers = [
        "来源系统", "来源库", "来源schema", "来源表英文名", "来源表中文名", "来源字段英文名",
        "来源字段中文名", "处理逻辑", "处理逻辑类型", "技术口径确认人", "备注",
    ]
    next_column = 12
    next_column = _add_group(sheet, next_column, "业务口径-借记卡", business_headers)
    next_column = _add_group(sheet, next_column, "溯源-借记卡系统", technical_headers)
    next_column = _add_group(sheet, next_column, "业务口径-信用卡", business_headers)
    _add_group(sheet, next_column, "溯源-贷款产品系统", technical_headers)

    fixed_values = [
        "CARD_PRODUCT_ID", "卡产品编号", "基础信息", "VARCHAR(32)", "银行卡产品唯一标识",
        "区分不同卡产品", "银行卡信息表", "卡产品编号", "EAST_CARD_PRODUCT_ID", "行内卡产品编号", "脱敏模拟数据",
    ]
    for column, value in enumerate(fixed_values, start=1):
        sheet.cell(4, column, value)

    _set_group_values(sheet, 4, 12, ["借记卡产品唯一编号", "是", "否", "否", "否", "银行卡部", "已调研"])
    _set_group_values(sheet, 4, 19, ["借记卡系统", "CARD_DB", "ODS", "CPS_CARDPRODUCT", "卡产品表", "CARD_PRODUCT_ID", "卡产品编号", "源字段直接取值", "direct", "科技部", "已确认"])
    _set_group_values(sheet, 4, 30, ["信用卡产品编号", "否", "否", "否", "否", "信用卡部", "待技术确认"])
    _set_group_values(sheet, 4, 37, ["信贷系统", "LOAN_DB", "ODS", "LOAN_PRODUCT", "贷款产品表", "PRODUCT_ID", "产品编号", "按产品号取值", "direct", "信贷科技", "仅技术溯源"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _add_group(sheet, start_column: int, title: str, headers: list[str]) -> int:
    end_column = start_column + len(headers) - 1
    sheet.cell(2, start_column, title)
    sheet.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=end_column)
    for offset, value in enumerate(headers):
        sheet.cell(3, start_column + offset, value)
    return end_column + 1


def _set_group_values(sheet, row: int, start_column: int, values: list[str]) -> None:
    for offset, value in enumerate(values):
        sheet.cell(row, start_column + offset, value)


@contextmanager
def _client() -> Iterator[TestClient]:
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)

    def override_get_db() -> Iterator[Session]:
        session = factory()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        with TestClient(app) as client:
            yield client
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(engine)


def _post(client: TestClient, path: str, payload: dict) -> dict:
    response = client.post(path, json=payload)
    response.raise_for_status()
    return response.json()


def _get(client: TestClient, path: str) -> dict | list[dict]:
    response = client.get(path)
    response.raise_for_status()
    return response.json()


def _put(client: TestClient, path: str, payload: dict) -> dict:
    response = client.put(path, json=payload)
    response.raise_for_status()
    return response.json()
