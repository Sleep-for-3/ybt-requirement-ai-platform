from collections.abc import Iterator
from contextlib import contextmanager
from io import BytesIO
import json
import zipfile
import subprocess
import sys
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.core.settings import get_settings
from app.main import app


def test_builtin_uat_suites_are_initialized_idempotently_in_display_order() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "示例 UAT 项目"})

        first = client.get(f"/api/projects/{project['id']}/uat-suites")
        second = client.get(f"/api/projects/{project['id']}/uat-suites")

        assert first.status_code == second.status_code == 200
        suites = first.json()
        assert len(suites) == 8
        assert [suite["suite_type"] for suite in suites] == [
            "end_to_end_delivery",
            "knowledge_and_citation",
            "catalog_and_source",
            "governance_workflow",
            "sql_lineage",
            "excel_fidelity",
            "permission_security",
            "deployment_readiness",
        ]
        assert second.json() == suites
        assert all(suite["is_system"] for suite in suites)
        assert all(
            [case["display_order"] for case in suite["cases"]]
            == sorted(case["display_order"] for case in suite["cases"])
            for suite in suites
        )


def test_builtin_automatic_cases_execute_registered_checks_instead_of_blocking() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "内置 UAT 执行项目"})
        suites = client.get(f"/api/projects/{project['id']}/uat-suites").json()
        delivery = next(item for item in suites if item["suite_type"] == "end_to_end_delivery")
        run = _post(client, f"/api/uat-suites/{delivery['id']}/runs", {"run_name": "内置自动检查"})

        executed = _post(client, f"/api/uat-runs/{run['id']}/execute", {})["run"]

        assert executed["status"] == "failed"
        assert {item["status"] for item in executed["results"]} == {"failed"}
        assert all(item["evidence_json"]["check_key"].startswith("end_to_end_delivery:") for item in executed["results"])


def test_uat_run_executes_cases_independently_and_retry_reuses_results() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "UAT 执行项目"})
        suite = _post(client, f"/api/projects/{project['id']}/uat-suites", {
            "suite_name": "执行语义套件",
            "suite_type": "custom",
            "description": "公开模拟执行检查",
            "cases": [
                _case("AUTO-PASS-1", "自动通过一", "automatic", "always_pass", "critical", 1),
                _case("AUTO-FAIL", "自动可控失败", "automatic", "always_fail", "high", 2),
                _case("AUTO-PASS-2", "失败后仍执行", "automatic", "always_pass", "medium", 3),
                _case("MANUAL", "人工确认", "manual", "manual", "medium", 4),
                _case("HYBRID", "混合确认", "hybrid", "always_pass", "medium", 5),
            ],
        })
        run = _post(client, f"/api/uat-suites/{suite['id']}/runs", {
            "run_name": "第一轮 UAT",
            "environment_name": "sanitized-test",
            "application_version": "test-version",
            "git_commit_sha": "a" * 40,
        })

        execution_response = client.post(
            f"/api/uat-runs/{run['id']}/execute",
            json={},
            headers={"X-Request-ID": "uat-execution-request"},
        )
        assert execution_response.status_code == 200, execution_response.text
        executed = execution_response.json()
        assert executed["run"]["status"] == "failed"
        assert [item["status"] for item in executed["run"]["results"]] == [
            "passed", "failed", "passed", "pending", "pending",
        ]
        assert executed["job"]["status"] == "partially_completed"
        assert executed["job"]["correlation_id"] == "uat-execution-request"
        result_ids = [item["id"] for item in executed["run"]["results"]]

        retried = _post(client, f"/api/uat-runs/{run['id']}/retry-failed", {})
        assert [item["id"] for item in retried["run"]["results"]] == result_ids
        assert len(retried["run"]["results"]) == 5
        assert retried["run"]["summary_json"]["attempt"] == 2


def test_controlled_failure_passes_on_retry_and_reuses_result() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "UAT 修复重跑项目"})
        suite = _post(client, f"/api/projects/{project['id']}/uat-suites", {
            "suite_name": "受控修复套件",
            "suite_type": "custom",
            "cases": [_case("RETRY", "首次失败修复后通过", "automatic", "fail_once_then_pass", "critical", 1)],
        })
        run = _post(client, f"/api/uat-suites/{suite['id']}/runs", {"run_name": "受控修复轮次"})

        first = _post(client, f"/api/uat-runs/{run['id']}/execute", {})
        assert first["run"]["status"] == "failed"
        first_result = first["run"]["results"][0]
        assert first_result["status"] == "failed"
        assert first_result["evidence_json"]["attempt"] == 1

        retried = _post(client, f"/api/uat-runs/{run['id']}/retry-failed", {})
        retry_result = retried["run"]["results"][0]
        assert retried["run"]["status"] == "passed"
        assert retried["run"]["summary_json"]["attempt"] == 2
        assert retry_result["id"] == first_result["id"]
        assert retry_result["status"] == "passed"
        assert retry_result["evidence_json"]["attempt"] == 2


def test_critical_finding_blocks_signoff_until_resolution_and_verification() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "UAT 签署项目"})
        suite = _post(client, f"/api/projects/{project['id']}/uat-suites", {
            "suite_name": "签署套件", "suite_type": "custom",
            "cases": [_case("PASS", "通过条件", "automatic", "always_pass", "critical", 1)],
        })
        run = _post(client, f"/api/uat-suites/{suite['id']}/runs", {"run_name": "签署轮次"})
        run = _post(client, f"/api/uat-runs/{run['id']}/execute", {})["run"]
        assert run["status"] == "passed"
        finding = _post(client, f"/api/uat-runs/{run['id']}/findings", {
            "uat_case_result_id": run["results"][0]["id"],
            "finding_type": "functional", "severity": "critical", "title": "受控阻断问题",
            "description": "使用完全虚构条件验证签署阻断。",
            "expected_behavior": "满足验收要求", "actual_behavior": "模拟条件未满足",
        })
        assert client.get(f"/api/uat-findings/{finding['id']}").json()["title"] == "受控阻断问题"
        blocked = client.post(f"/api/uat-runs/{run['id']}/signoff", json={
            "signoff_role": "business_owner", "signoff_status": "approved", "comment": "不应通过",
        })
        assert blocked.status_code == 409
        rejected_finding = client.patch(f"/api/uat-findings/{finding['id']}", json={"status": "rejected"})
        assert rejected_finding.status_code == 200
        still_blocked = client.post(f"/api/uat-runs/{run['id']}/signoff", json={
            "signoff_role": "business_owner", "signoff_status": "approved", "comment": "拒绝问题不能绕过阻断",
        })
        assert still_blocked.status_code == 409

        assigned = client.patch(f"/api/uat-findings/{finding['id']}", json={"assigned_role": "technical_analyst", "status": "assigned"})
        assert assigned.status_code == 200
        resolved = _post(client, f"/api/uat-findings/{finding['id']}/resolve", {"resolution_text": "已修复模拟条件并完成回归。"})
        assert resolved["status"] == "resolved"
        verified = _post(client, f"/api/uat-findings/{finding['id']}/verify", {"verification_comment": "脱敏回归通过。"})
        assert verified["status"] == "verified"

        for role in ("business_owner", "technical_owner", "project_manager", "final_acceptance"):
            _post(client, f"/api/uat-runs/{run['id']}/signoff", {
                "signoff_role": role, "signoff_status": "approved", "comment": f"{role} 脱敏验收通过",
            })
        repeated = _post(client, f"/api/uat-runs/{run['id']}/signoff", {
            "signoff_role": "business_owner", "signoff_status": "rejected", "comment": "保留历史的新记录",
        })
        signoffs = client.get(f"/api/uat-runs/{run['id']}/signoffs").json()
        assert len(signoffs) == 5
        assert repeated["id"] != signoffs[0]["id"]
        assert [item["signoff_status"] for item in signoffs if item["signoff_role"] == "business_owner"] == ["approved", "rejected"]


def test_uat_pack_rejects_path_traversal_and_exports_sanitized_evidence(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_DIR", str(tmp_path / "storage"))
    from app.services.storage import get_storage_service
    get_storage_service.cache_clear()
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "材料和报告项目"})
        safe_zip = _zip_bytes({
            "一表通目标字段模板.xlsx": b"DEMO target template",
            "银行正式交付模板.xlsx": b"DEMO delivery template",
            "历史业务口径.xlsx": b"DEMO business caliber",
            "历史技术溯源.xlsx": b"DEMO technical lineage",
            "监管答疑.xlsx": b"DEMO regulatory qa",
            "数据字典.xlsx": b"DEMO dictionary",
            "load_customer_v1.sql": b"-- DEMO SQL, ingestion evidence only",
            "run_customer.sh": b"#!/bin/sh\n# DEMO script, never executed",
            "README.md": b"# DEMO UAT instructions",
        })
        uploaded = client.post(f"/api/projects/{project['id']}/uat-packs/upload", data={"pack_name": "公开模拟材料"}, files=[("files", ("demo.zip", safe_zip, "application/zip"))])
        assert uploaded.status_code == 201, uploaded.text
        pack = uploaded.json()
        assert len(pack["items"]) == 9
        validation = _post(client, f"/api/uat-packs/{pack['id']}/validate", {})
        assert validation["valid"] is True

        unsafe = client.post(f"/api/projects/{project['id']}/uat-packs/upload", files=[("files", ("unsafe.zip", _zip_bytes({"../escape.sql": b"SELECT 1"}), "application/zip"))])
        assert unsafe.status_code == 400
        executable = client.post(f"/api/projects/{project['id']}/uat-packs/upload", files=[("files", ("malware.exe", b"MZ-demo", "application/octet-stream"))])
        assert executable.status_code == 400

        suite = _post(client, f"/api/projects/{project['id']}/uat-suites", {"suite_name": "报告套件", "suite_type": "custom", "cases": [_case("PASS", "报告通过", "automatic", "always_pass", "critical", 1)]})
        run = _post(client, f"/api/uat-suites/{suite['id']}/runs", {"run_name": "报告轮次", "git_commit_sha": "b" * 40})
        run = _post(client, f"/api/uat-runs/{run['id']}/execute", {})["run"]
        report = client.get(f"/api/uat-runs/{run['id']}/report")
        assert report.status_code == 200
        workbook = load_workbook(BytesIO(report.content))
        assert workbook.sheetnames == ["验收概览", "测试套件", "测试案例", "失败案例", "阻断案例", "问题清单", "修复记录", "签署记录", "环境信息", "版本信息"]
        evidence = client.get(f"/api/uat-runs/{run['id']}/evidence-package")
        assert evidence.status_code == 200
        with zipfile.ZipFile(BytesIO(evidence.content)) as archive:
            names = archive.namelist()
            assert "SHA256SUMS" in names
            assert "uat-report.xlsx" in names
            assert not any(".env" in name or "token" in name.lower() or "database" in name.lower() for name in names)
            assert all(not name.startswith("/") and ".." not in name for name in names)
            version = json.loads(archive.read("version.json"))
            assert version["alembic_head_revision"] == "202607230012"
            assert version["alembic_revision"] is None


def test_uat_pack_enforces_combined_upload_limit_before_persistence(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("UAT_ZIP_MAX_TOTAL_BYTES", "10")
    monkeypatch.setenv("STORAGE_DIR", str(tmp_path / "aggregate-storage"))
    get_settings.cache_clear()
    from app.services.storage import get_storage_service
    get_storage_service.cache_clear()
    try:
        with _client() as client:
            project = _post(client, "/api/projects", {"name": "整包容量限制项目"})
            response = client.post(
                f"/api/projects/{project['id']}/uat-packs/upload",
                files=[
                    ("files", ("one.sql", b"123456", "text/plain")),
                    ("files", ("two.sql", b"123456", "text/plain")),
                ],
            )
            assert response.status_code == 400
            assert "total size limit" in response.text
            assert client.get(f"/api/projects/{project['id']}/uat-packs").json() == []
    finally:
        get_settings.cache_clear()
        get_storage_service.cache_clear()


def test_uat_report_escapes_formula_like_user_content() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "公式注入防护项目"})
        suite = _post(client, f"/api/projects/{project['id']}/uat-suites", {
            "suite_name": "=HYPERLINK(\"https://example.invalid\",\"unsafe\")",
            "suite_type": "custom",
            "cases": [_case("PASS", "+SUM(1,1)", "automatic", "always_pass", "medium", 1)],
        })
        run = _post(client, f"/api/uat-suites/{suite['id']}/runs", {"run_name": "@formula-like-run"})
        _post(client, f"/api/uat-runs/{run['id']}/execute", {})

        report = client.get(f"/api/uat-runs/{run['id']}/report")
        workbook = load_workbook(BytesIO(report.content), data_only=False)

        assert workbook["验收概览"]["C2"].value == "'@formula-like-run"
        assert workbook["验收概览"]["C2"].data_type == "s"
        assert workbook["测试套件"]["B2"].value.startswith("'=")
        assert workbook["测试案例"]["B2"].value.startswith("'+")


def test_uat_permissions_hide_cross_institution_resources_and_block_viewer_execution() -> None:
    with _client() as client:
        password = "test-only-platform-admin-password"
        bootstrap = client.post("/api/admin/bootstrap", json={
            "institution_code": "PLATFORM",
            "institution_name": "公开测试平台",
            "institution_type": "platform_operator",
            "username": "uat_platform_admin",
            "display_name": "UAT 平台管理员",
            "email": "uat-platform-admin@example.invalid",
            "password": password,
        })
        assert bootstrap.status_code == 201, bootstrap.text
        admin_token = _post(client, "/api/auth/login", {"username": "uat_platform_admin", "password": password})["access_token"]
        admin_headers = {"Authorization": f"Bearer {admin_token}"}
        bank = client.post("/api/admin/institutions", headers=admin_headers, json={
            "institution_code": "UAT_BANK",
            "institution_name": "UAT 示例银行",
            "institution_type": "bank",
        })
        assert bank.status_code == 201, bank.text
        other_bank = client.post("/api/admin/institutions", headers=admin_headers, json={
            "institution_code": "OTHER_UAT_BANK",
            "institution_name": "其他 UAT 示例银行",
            "institution_type": "bank",
        })
        assert other_bank.status_code == 201, other_bank.text
        project = client.post("/api/projects", headers=admin_headers, json={
            "name": "UAT 权限隔离项目",
            "institution_id": bank.json()["id"],
        })
        assert project.status_code == 200, project.text
        suite = client.post(f"/api/projects/{project.json()['id']}/uat-suites", headers=admin_headers, json={
            "suite_name": "权限验证套件",
            "suite_type": "custom",
            "cases": [_case("PASS", "只读验证", "automatic", "always_pass", "medium", 1)],
        })
        assert suite.status_code == 201, suite.text
        run = client.post(f"/api/uat-suites/{suite.json()['id']}/runs", headers=admin_headers, json={"run_name": "权限验证轮次"})
        assert run.status_code == 201, run.text

        viewer_password = "test-only-viewer-password"
        viewer = client.post("/api/admin/users", headers=admin_headers, json={
            "username": "uat_viewer",
            "display_name": "UAT 只读用户",
            "email": "uat-viewer@example.invalid",
            "password": viewer_password,
            "institution_id": bank.json()["id"],
            "institution_role": "member",
        })
        assert viewer.status_code == 201, viewer.text
        membership = client.post(f"/api/projects/{project.json()['id']}/members", headers=admin_headers, json={
            "user_id": viewer.json()["id"],
            "project_role": "viewer",
        })
        assert membership.status_code == 201, membership.text
        viewer_token = _post(client, "/api/auth/login", {"username": "uat_viewer", "password": viewer_password})["access_token"]
        viewer_headers = {"Authorization": f"Bearer {viewer_token}"}
        assert client.get(f"/api/uat-suites/{suite.json()['id']}", headers=viewer_headers).status_code == 200
        assert client.post(f"/api/uat-runs/{run.json()['id']}/execute", headers=viewer_headers, json={}).status_code == 403

        outsider_password = "test-only-outsider-password"
        outsider = client.post("/api/admin/users", headers=admin_headers, json={
            "username": "uat_outsider",
            "display_name": "UAT 跨机构用户",
            "email": "uat-outsider@example.invalid",
            "password": outsider_password,
            "institution_id": other_bank.json()["id"],
            "institution_role": "member",
        })
        assert outsider.status_code == 201, outsider.text
        outsider_token = _post(client, "/api/auth/login", {"username": "uat_outsider", "password": outsider_password})["access_token"]
        outsider_headers = {"Authorization": f"Bearer {outsider_token}"}
        assert client.get(f"/api/uat-suites/{suite.json()['id']}", headers=outsider_headers).status_code == 404


def test_critical_precondition_blocks_dependents_manual_completion_and_cancel() -> None:
    with _client() as client:
        project = _post(client, "/api/projects", {"name": "UAT 阻断项目"})
        suite = _post(client, f"/api/projects/{project['id']}/uat-suites", {
            "suite_name": "阻断和人工套件", "suite_type": "custom", "cases": [
                _case("CRITICAL", "关键前置", "automatic", "always_fail", "critical", 1),
                {**_case("DEPENDENT", "依赖项", "automatic", "always_pass", "high", 2), "precondition_json": {"check_key": "always_pass", "depends_on": ["CRITICAL"]}},
                _case("INDEPENDENT", "独立项", "automatic", "always_pass", "medium", 3),
                _case("MANUAL", "人工项", "manual", "manual", "medium", 4),
            ],
        })
        clone = _post(client, f"/api/uat-suites/{suite['id']}/clone", {"suite_name": "阻断套件副本"})
        assert len(clone["cases"]) == 4 and clone["is_system"] is False
        run = _post(client, f"/api/uat-suites/{suite['id']}/runs", {"run_name": "阻断轮次"})
        executed = _post(client, f"/api/uat-runs/{run['id']}/execute", {})["run"]
        assert [item["status"] for item in executed["results"]] == ["failed", "blocked", "passed", "pending"]
        manual = executed["results"][3]
        completed = _post(client, f"/api/uat-case-results/{manual['id']}/complete-manual", {"status": "passed", "actual_result_json": {"confirmed": True}, "evidence_json": {"reference": "SANITIZED-UAT-001"}})
        assert completed["status"] == "passed"
        attached = _post(client, f"/api/uat-case-results/{manual['id']}/attach-evidence", {"evidence": {"password": "must-not-persist", "note": "脱敏证据"}})
        assert "password" not in attached["evidence_json"] and attached["evidence_json"]["note"] == "脱敏证据"

        draft = _post(client, f"/api/uat-suites/{suite['id']}/runs", {"run_name": "取消轮次"})
        cancelled = _post(client, f"/api/uat-runs/{draft['id']}/cancel", {})
        assert cancelled["status"] == "cancelled"


def test_demo_uat_pack_generator_is_repeatable(tmp_path) -> None:
    output = tmp_path / "demo-pack"
    script = Path(__file__).resolve().parents[2] / "scripts" / "generate_demo_uat_pack.py"
    first = subprocess.run([sys.executable, str(script), "--output", str(output)], capture_output=True, text=True, timeout=60)
    assert first.returncode == 0, first.stdout + first.stderr
    first_hashes = {item.name: item.read_bytes() for item in output.iterdir()}
    second = subprocess.run([sys.executable, str(script), "--output", str(output)], capture_output=True, text=True, timeout=60)
    assert second.returncode == 0, second.stdout + second.stderr
    assert {item.name: item.read_bytes() for item in output.iterdir()} == first_hashes
    assert set(first_hashes) == {"一表通目标字段模板.xlsx", "银行正式交付模板.xlsx", "历史业务口径.xlsx", "历史技术溯源.xlsx", "监管答疑.xlsx", "数据字典.xlsx", "load_customer_v1.sql", "load_customer_v2.sql", "run_customer.sh", "expected_manifest.json", "README.md"}


@contextmanager
def _client() -> Iterator[TestClient]:
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, autoflush=False)

    def override() -> Iterator[Session]:
        session = factory()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = override
    try:
        with TestClient(app) as client:
            yield client
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(engine)


def _post(client: TestClient, path: str, payload: dict) -> dict:
    response = client.post(path, json=payload)
    assert response.status_code in {200, 201}, response.text
    return response.json()


def _case(code: str, name: str, mode: str, check_key: str, severity: str, order: int) -> dict:
    return {
        "case_code": code,
        "case_name": name,
        "case_category": "custom",
        "precondition_json": {"check_key": check_key},
        "input_requirement_json": {"sanitized_fixture_only": True},
        "expected_result_json": {"status": "passed"},
        "execution_mode": mode,
        "severity": severity,
        "display_order": order,
    }


def _zip_bytes(files: dict[str, bytes]) -> bytes:
    stream = BytesIO()
    with zipfile.ZipFile(stream, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)
    return stream.getvalue()
