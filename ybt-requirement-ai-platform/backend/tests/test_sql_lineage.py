from collections.abc import Iterator
from contextlib import contextmanager
import base64
from pathlib import Path
from io import BytesIO
from types import SimpleNamespace
from zipfile import ZipFile
from openpyxl import load_workbook

from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.main import app
from app.models import (
    BackgroundJob, CatalogColumn, CatalogSchema, CatalogTable, CodeRepository, DataSource, Institution, LineageEdge, LineageNode,
    MartField, MartTable, MartToYbtMapping, ProductScenario, Project, ReviewTask,
    ScenarioBusinessMapping, ScenarioTechnicalLineage, ScriptChangeSet, ScriptFile as ScriptFileModel, ScriptFileVersion,
    SourceToMartMapping, TargetField, TargetTable, TemplateVariable, User,
)
from app.services.lineage.resolver import resolve_lineage_node
from app.services.lineage.impact_analyzer import persist_change_impact
from app.services.governance.audit import redact_summary
from app.services.lineage.preprocessing import preprocess_sql
from app.services.lineage.sql_parser import parse_sql_lineage
from app.services.lineage.shell_parser import parse_shell_dependencies
from app.services.lineage.version_diff import compare_shell_versions, compare_sql_versions
from app.services.lineage.archive_ingestion import read_safe_script_archive
from app.services.lineage.archive_ingestion import ArchivedScript
from app.services.lineage.git_repository import GitRepositorySnapshot, read_git_repository_scripts
from app.services.lineage.ingestion import ScriptIngestionService
from app.services.lineage.exporter import export_lineage_workbook
from app.services.storage.local import LocalStorageService


def test_template_variables_are_preserved_and_instance_table_is_parseable() -> None:
    source = """
    insert into #{INSTANCE_GB}.MART_CUSTOMER (CERT_TYPE)
    select CERT_TYPE from ${SCHEMA}.ECIF_CUSTOMER where BIZ_DATE = :biz_date
    """

    result = preprocess_sql(source, variables={"INSTANCE_GB": "SAFE_SCHEMA"})

    assert "#{INSTANCE_GB}.MART_CUSTOMER" in result.original_sql
    assert "SAFE_SCHEMA.MART_CUSTOMER" in result.parse_sql
    assert "SAFE_SCHEMA.ECIF_CUSTOMER" in result.parse_sql
    assert ":biz_date" in result.normalized_sql
    assert {item.expression for item in result.variables} == {
        "#{INSTANCE_GB}",
        "${SCHEMA}",
        ":biz_date",
    }
    assert any("SCHEMA" in warning for warning in result.warnings)


def test_insert_select_produces_table_and_column_lineage() -> None:
    sql = """
    insert into MART.MART_CUSTOMER (CERT_TYPE, CUSTOMER_COUNT)
    select c.CERT_TYPE, count(c.CUSTOMER_ID)
    from ODS.ECIF_CUSTOMER c
    where c.ENABLED_FLAG = 'Y'
    group by c.CERT_TYPE
    """

    result = parse_sql_lineage(sql, dialect="sqlite")

    assert result.parse_status == "parsed"
    assert any(
        edge.source.logical_name == "ODS.ECIF_CUSTOMER"
        and edge.target.logical_name == "MART.MART_CUSTOMER"
        and edge.edge_type == "reads_from"
        for edge in result.edges
    )
    cert_edge = next(
        edge for edge in result.edges
        if edge.target.logical_name == "MART.MART_CUSTOMER.CERT_TYPE"
        and edge.edge_type == "derives_from"
    )
    assert cert_edge.source.logical_name == "ODS.ECIF_CUSTOMER.CERT_TYPE"
    assert "ENABLED_FLAG" in (cert_edge.filter_condition or "")
    count_edge = next(
        edge for edge in result.edges
        if edge.target.logical_name == "MART.MART_CUSTOMER.CUSTOMER_COUNT"
        and edge.edge_type == "aggregates"
    )
    assert count_edge.source.logical_name == "ODS.ECIF_CUSTOMER.CUSTOMER_ID"
    assert "COUNT" in count_edge.transformation_expression.upper()


def test_multi_statement_sql_retains_successful_lineage_when_one_statement_fails() -> None:
    result = parse_sql_lineage("""
        insert into MART.T1 (A) select A from ODS.S1;
        select from where definitely invalid ???;
        insert into MART.T2 (B) select B from ODS.S2;
    """, dialect="sqlite")

    assert result.parse_status == "partially_parsed"
    assert [item.parse_status for item in result.statements] == ["parsed", "failed", "parsed"]
    assert {edge.target.logical_name for edge in result.edges if edge.target.node_type == "column"} == {
        "MART.T1.A",
        "MART.T2.B",
    }
    assert all(edge.confidence_level == "low" for edge in result.edges)
    assert any("Statement 2 parse failed" in warning for warning in result.warnings)


def test_script_ingestion_persists_discovered_template_variables(tmp_path: Path, db_session: Session) -> None:
    institution = Institution(institution_code="template-bank", institution_name="Template Bank")
    db_session.add(institution)
    db_session.flush()
    project = Project(name="template discovery", institution_id=institution.id)
    db_session.add(project)
    db_session.flush()

    ScriptIngestionService(db_session, LocalStorageService(tmp_path / "storage")).ingest(
        project=project,
        data=b"insert into #{INSTANCE_GB}.TARGET(A) select A from ${SCHEMA}.SOURCE where D=:biz_date",
        file_name="template.sql",
        relative_path="template.sql",
        dialect="sqlite",
        actor_user_id=None,
    )

    rows = db_session.query(TemplateVariable).filter(TemplateVariable.project_id == project.id).all()
    assert {(row.variable_name, row.variable_type, row.confirmed) for row in rows} == {
        ("INSTANCE_GB", "identifier_or_value", False),
        ("SCHEMA", "identifier_or_value", False),
        ("biz_date", "identifier_or_value", False),
    }


def test_manual_sql_upload_is_version_idempotent_and_queryable(tmp_path: Path, monkeypatch) -> None:
    import app.api.lineage as lineage_api

    storage = LocalStorageService(tmp_path / "storage")
    monkeypatch.setattr(lineage_api, "get_storage_service", lambda: storage)
    sql = b"insert into MART.TARGET (CERT_TYPE) select CERT_TYPE from ODS.CUSTOMER"

    with _client() as client:
        project = client.post("/api/projects", json={"name": "lineage", "institution_id": 1}).json()
        first = client.post(
            f"/api/projects/{project['id']}/scripts/upload",
            data={"relative_path": "sql/load_customer.sql", "dialect": "sqlite"},
            files={"file": ("load_customer.sql", sql, "text/plain")},
        )
        assert first.status_code == 200, first.text
        second = client.post(
            f"/api/projects/{project['id']}/scripts/upload",
            data={"relative_path": "sql/load_customer.sql", "dialect": "sqlite"},
            files={"file": ("load_customer.sql", sql, "text/plain")},
        )
        assert second.status_code == 200, second.text
        assert second.json()["script_file_id"] == first.json()["script_file_id"]
        assert second.json()["version_id"] == first.json()["version_id"]
        assert second.json()["deduplicated"] is True

        graph = client.get(f"/api/projects/{project['id']}/lineage/graph?direction=both&depth=3")
        assert graph.status_code == 200, graph.text
        assert {node["logical_name"] for node in graph.json()["nodes"]} >= {
            "ODS.CUSTOMER.CERT_TYPE",
            "MART.TARGET.CERT_TYPE",
        }
        assert storage.read(first.json()["storage_key"]) == sql


def test_shell_parser_detects_scripts_sql_clients_and_arguments_without_execution() -> None:
    source = """
    SQL_FILE=sql/load_customer.sql
    sh ybt_04_loadData_HYF.sh HYF_TABLE YBT
    source common/env.sh
    psql -f ${SQL_FILE}
    mysql reporting < sql/check_customer.sql
    """

    result = parse_shell_dependencies(source)

    assert [(item.dependency_type, item.target_path) for item in result.dependencies] == [
        ("calls_script", "ybt_04_loadData_HYF.sh"),
        ("sources_script", "common/env.sh"),
        ("executes_sql", "sql/load_customer.sql"),
        ("executes_sql", "sql/check_customer.sql"),
    ]
    assert result.dependencies[0].arguments == ("HYF_TABLE", "YBT")
    assert all(item.source_line_start == item.source_line_end for item in result.dependencies)


def test_uploaded_shell_dependency_resolves_to_uploaded_sql(tmp_path: Path, monkeypatch) -> None:
    import app.api.lineage as lineage_api

    monkeypatch.setattr(lineage_api, "get_storage_service", lambda: LocalStorageService(tmp_path / "storage"))
    with _client() as client:
        project = client.post("/api/projects", json={"name": "shell lineage", "institution_id": 1}).json()
        sql = client.post(
            f"/api/projects/{project['id']}/scripts/upload",
            data={"relative_path": "sql/load.sql", "dialect": "sqlite"},
            files={"file": ("load.sql", b"insert into T.B select C from S.A", "text/plain")},
        ).json()
        shell = client.post(
            f"/api/projects/{project['id']}/scripts/upload",
            data={"relative_path": "run.sh"},
            files={"file": ("run.sh", b"psql -f sql/load.sql\n", "text/plain")},
        )
        assert shell.status_code == 200, shell.text
        detail = client.get(f"/api/scripts/{shell.json()['script_file_id']}")
        assert detail.status_code == 200, detail.text
        dependency = detail.json()["dependencies"][0]
        assert dependency["dependency_type"] == "executes_sql"
        assert dependency["child_script_file_id"] == sql["script_file_id"]
        assert dependency["source_line_start"] == 1


def test_cte_column_lineage_resolves_to_physical_source() -> None:
    result = parse_sql_lineage("""
        insert into MART.CUSTOMER_SUMMARY (CERT_TYPE)
        with scoped as (
            select c.CERT_TYPE from ODS.ECIF_CUSTOMER c where c.STATUS = 'A'
        )
        select scoped.CERT_TYPE from scoped
    """, dialect="sqlite")

    column_edges = [edge for edge in result.edges if edge.target.logical_name.endswith(".CERT_TYPE")]
    assert any(edge.source.logical_name == "ODS.ECIF_CUSTOMER.CERT_TYPE" for edge in column_edges)
    assert all(edge.source.logical_name != "scoped.CERT_TYPE" for edge in column_edges)


def test_comment_and_format_only_change_is_non_semantic() -> None:
    old = "insert into T.B (C) select C from S.A"
    new = """-- explain the load
        insert  into T.B (C)
        select C
        from S.A
    """

    result = compare_sql_versions(old, new, dialect="sqlite")

    assert result.semantic_changed is False
    assert [item.change_category for item in result.items] == ["non_semantic"]
    assert result.severity == "low"


def test_source_removal_and_filter_change_are_classified_by_risk() -> None:
    old = """
        insert into MART.T (A, B)
        select s.A, s.B from ODS.S s where s.STATUS = 'A'
    """
    new = """
        insert into MART.T (A)
        select s.A from ODS.S s where s.STATUS = 'B'
    """

    result = compare_sql_versions(old, new, dialect="sqlite")

    categories = {item.change_category for item in result.items}
    assert "source_column_removed" in categories
    assert "filter_changed" in categories
    assert result.severity == "critical"


def test_catalog_resolution_auto_binds_only_a_unique_candidate(db_session: Session) -> None:
    project = Project(name="resolve")
    db_session.add(project); db_session.flush()
    datasource = DataSource(project_id=project.id, name="warehouse", db_type="sqlite")
    db_session.add(datasource); db_session.flush()
    schema = CatalogSchema(project_id=project.id, datasource_id=datasource.id, schema_name="ODS")
    db_session.add(schema); db_session.flush()
    table = CatalogTable(project_id=project.id, datasource_id=datasource.id, catalog_schema_id=schema.id, schema_name="ODS", table_name="CUSTOMER")
    db_session.add(table); db_session.flush()
    column = CatalogColumn(project_id=project.id, datasource_id=datasource.id, catalog_table_id=table.id, schema_name="ODS", table_name="CUSTOMER", column_name="CERT_TYPE", ordinal_position=1)
    unique_node = LineageNode(project_id=project.id, node_type="column", logical_name="ODS.CUSTOMER.CERT_TYPE", schema_name="ODS", table_name="CUSTOMER", column_name="CERT_TYPE")
    db_session.add_all([column, unique_node]); db_session.flush()

    unique = resolve_lineage_node(db_session, unique_node)
    assert unique_node.catalog_column_id == column.id
    assert unique_node.unresolved_flag is False
    assert unique.candidates == ()

    other_schema = CatalogSchema(project_id=project.id, datasource_id=datasource.id, schema_name="DWD")
    db_session.add(other_schema); db_session.flush()
    other_table = CatalogTable(project_id=project.id, datasource_id=datasource.id, catalog_schema_id=other_schema.id, schema_name="DWD", table_name="CUSTOMER")
    db_session.add(other_table); db_session.flush()
    other_column = CatalogColumn(project_id=project.id, datasource_id=datasource.id, catalog_table_id=other_table.id, schema_name="DWD", table_name="CUSTOMER", column_name="CERT_TYPE", ordinal_position=1)
    ambiguous_node = LineageNode(project_id=project.id, node_type="column", logical_name="CUSTOMER.CERT_TYPE", table_name="CUSTOMER", column_name="CERT_TYPE")
    db_session.add_all([other_column, ambiguous_node]); db_session.flush()

    ambiguous = resolve_lineage_node(db_session, ambiguous_node)
    assert ambiguous_node.catalog_column_id is None
    assert ambiguous_node.unresolved_flag is True
    assert len(ambiguous.candidates) == 2


def test_second_script_version_creates_change_set_and_impact(tmp_path: Path, monkeypatch) -> None:
    import app.api.lineage as lineage_api

    monkeypatch.setattr(lineage_api, "get_storage_service", lambda: LocalStorageService(tmp_path / "storage"))
    with _client() as client:
        project = client.post("/api/projects", json={"name": "impact", "institution_id": 1}).json()
        first = b"insert into MART.T (A, B) select A, B from ODS.S where STATUS = 'A'"
        second = b"insert into MART.T (A) select A from ODS.S where STATUS = 'B'"
        for content in (first, second):
            response = client.post(
                f"/api/projects/{project['id']}/scripts/upload",
                data={"relative_path": "load.sql", "dialect": "sqlite"},
                files={"file": ("load.sql", content, "text/plain")},
            )
            assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["version_no"] == 2
        assert payload["change_set_id"]
        assert payload["impact_id"]
        assert "source_column_removed" in payload["change_categories"]
        assert payload["impact_severity"] == "critical"

        change = client.get(f"/api/lineage/changes/{payload['change_set_id']}")
        assert change.status_code == 200, change.text
        assert change.json()["impact"]["severity"] == "critical"


def test_zip_reader_blocks_zip_slip_and_reads_only_safe_scripts() -> None:
    safe = BytesIO()
    with ZipFile(safe, "w") as archive:
        archive.writestr("sql/load.sql", "select A from T")
        archive.writestr("run.sh", "psql -f sql/load.sql")
    files = read_safe_script_archive(safe.getvalue())
    assert [item.relative_path for item in files] == ["sql/load.sql", "run.sh"]

    malicious = BytesIO()
    with ZipFile(malicious, "w") as archive:
        archive.writestr("../../escape.sql", "select 1")
    try:
        read_safe_script_archive(malicious.getvalue())
    except ValueError as exc:
        assert "unsafe path" in str(exc).lower()
    else:
        raise AssertionError("Zip Slip payload was accepted")


def test_zip_upload_runs_as_idempotent_retryable_background_job(tmp_path: Path, monkeypatch) -> None:
    import app.api.lineage as lineage_api
    import app.services.lineage.jobs as lineage_jobs

    storage = LocalStorageService(tmp_path / "storage")
    monkeypatch.setattr(lineage_api, "get_storage_service", lambda: storage)
    monkeypatch.setattr(lineage_jobs, "get_storage_service", lambda: storage)
    stream = BytesIO()
    with ZipFile(stream, "w") as archive:
        archive.writestr("load.sql", "insert into T.B(A) select A from S.C")
        archive.writestr("run.sh", "psql -f load.sql")
    with _client() as client:
        project = client.post("/api/projects", json={"name": "archive", "institution_id": 1}).json()
        responses = [client.post(
            f"/api/projects/{project['id']}/scripts/upload-zip",
            data={"dialect": "sqlite"}, files={"file": ("scripts.zip", stream.getvalue(), "application/zip")},
        ) for _ in range(2)]
        assert all(item.status_code == 200 for item in responses)
        first, repeated = (item.json() for item in responses)
        assert first["status"] == "completed"
        assert first["job_id"] == repeated["job_id"]
        assert len(first["items"]) == 2
        assert len(client.get(f"/api/projects/{project['id']}/scripts").json()) == 2


def test_git_reader_does_not_execute_checkout_hooks(tmp_path: Path) -> None:
    repository = tmp_path / "repo"
    repository.mkdir()
    subprocess.run(["git", "init", "-b", "main"], cwd=repository, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "test@example.invalid"], cwd=repository, check=True)
    subprocess.run(["git", "config", "user.name", "Lineage Test"], cwd=repository, check=True)
    (repository / "load.sql").write_text("select A from T", encoding="utf-8")
    hooks = repository / ".git" / "hooks"
    marker = tmp_path / "hook-ran"
    hook = hooks / "post-checkout"
    hook.write_text(f"#!/bin/sh\nprintf ran > '{marker.as_posix()}'\n", encoding="utf-8")
    hook.chmod(0o755)
    subprocess.run(["git", "add", "load.sql"], cwd=repository, check=True)
    subprocess.run(["git", "commit", "-m", "safe fixture"], cwd=repository, check=True, capture_output=True)

    result = read_git_repository_scripts(str(repository), branch="main")

    assert result.files[0].relative_path == "load.sql"
    assert result.files[0].content == b"select A from T"
    assert not marker.exists()


def test_code_repository_rejects_url_credentials_and_unapproved_hosts() -> None:
    with _client() as client:
        project = client.post("/api/projects", json={"name": "git security", "institution_id": 1}).json()
        endpoint = f"/api/projects/{project['id']}/code-repositories"
        for repository_url in (
            "https://github.com/example/repo.git?access_token=secret",
            "https://github.com/example/repo.git#token=secret",
            "https://user:secret@github.com/example/repo.git",
            "https://internal.example.invalid/example/repo.git",
        ):
            response = client.post(endpoint, json={"repository_name": repository_url, "repository_url": repository_url})
            assert response.status_code == 400, response.text


def test_code_repository_local_path_must_be_inside_an_allowed_root(tmp_path: Path, monkeypatch) -> None:
    import app.api.lineage as lineage_api

    allowed = tmp_path / "allowed"
    allowed.mkdir()
    inside = allowed / "repo.git"
    outside = tmp_path / "outside.git"
    monkeypatch.setattr(
        lineage_api,
        "get_settings",
        lambda: SimpleNamespace(
            lineage_git_allowed_host_list=["github.com"],
            lineage_git_allowed_local_root_list=[str(allowed)],
        ),
    )
    with _client() as client:
        project = client.post("/api/projects", json={"name": "local git security", "institution_id": 1}).json()
        endpoint = f"/api/projects/{project['id']}/code-repositories"
        accepted = client.post(endpoint, json={"repository_name": "inside", "repository_url": str(inside)})
        rejected = client.post(endpoint, json={"repository_name": "outside", "repository_url": str(outside)})

    assert accepted.status_code == 200, accepted.text
    assert rejected.status_code == 400, rejected.text


def test_git_credential_is_only_in_the_subprocess_environment(monkeypatch) -> None:
    import app.services.lineage.git_repository as git_repository

    secret = "token-with-sensitive-value"
    captured: dict = {}

    def fail_clone(arguments, **kwargs):
        captured["arguments"] = arguments
        captured["env"] = kwargs["env"]
        return SimpleNamespace(returncode=128, stdout=b"", stderr=f"remote rejected {secret}".encode())

    monkeypatch.setattr(git_repository.subprocess, "run", fail_clone)
    try:
        read_git_repository_scripts("https://github.com/example/private.git", credential=secret)
    except ValueError as exc:
        message = str(exc)
    else:
        raise AssertionError("A failed Git clone unexpectedly succeeded")

    assert any(secret.encode() in base64.b64decode(value.split()[-1]) for key, value in captured["env"].items() if key.startswith("GIT_CONFIG_VALUE_"))
    assert secret not in " ".join(captured["arguments"])
    assert secret not in message
    assert "private.git" not in message


def test_git_repository_detects_added_renamed_and_deleted_scripts(tmp_path: Path, db_session: Session, monkeypatch) -> None:
    import app.services.lineage.jobs as lineage_jobs

    institution = Institution(institution_code="git-change-bank", institution_name="Git Change Bank")
    user = User(username="git-change-owner", status="active")
    db_session.add_all([institution, user])
    db_session.flush()
    project = Project(name="git changes", institution_id=institution.id)
    db_session.add(project)
    db_session.flush()
    repository = CodeRepository(
        institution_id=institution.id,
        project_id=project.id,
        repository_name="local",
        repository_type="git_repository",
        repository_url=str(tmp_path),
        default_branch="main",
        enabled=True,
        created_by=user.id,
    )
    db_session.add(repository)
    db_session.flush()
    storage = LocalStorageService(tmp_path / "storage")
    content = b"insert into MART.T(A) select A from ODS.S"
    added = ScriptIngestionService(db_session, storage).ingest(
        project=project, data=content, file_name="old.sql", relative_path="old.sql",
        dialect="sqlite", actor_user_id=user.id, code_repository_id=repository.id,
    )
    assert added.change_set is not None and added.change_set.change_type == "added"

    renamed_count = lineage_jobs._detect_repository_renames(
        db_session,
        repository,
        (ArchivedScript("renamed.sql", "renamed.sql", content),),
        user.id,
    )
    db_session.flush()
    assert renamed_count == 1
    assert added.script_file.relative_path == "renamed.sql"
    assert db_session.scalar(select(ScriptChangeSet).where(
        ScriptChangeSet.script_file_id == added.script_file.id,
        ScriptChangeSet.change_type == "renamed",
    )) is not None

    job = BackgroundJob(
        institution_id=institution.id,
        project_id=project.id,
        idempotency_key="delete-sync",
        job_type="script_repository_sync",
        status="running",
        progress=1,
        payload_summary_json={"repository_id": repository.id},
        result_summary_json={},
        created_by=user.id,
    )
    db_session.add(job)
    db_session.commit()
    monkeypatch.setattr(
        lineage_jobs,
        "read_git_repository_scripts",
        lambda *args, **kwargs: GitRepositorySnapshot("a" * 40, ()),
    )
    monkeypatch.setattr(
        lineage_jobs,
        "get_settings",
        lambda: SimpleNamespace(
            lineage_git_allowed_host_list=["github.com"],
            lineage_git_allowed_local_root_list=[str(tmp_path)],
            lineage_repository_max_bytes=1024,
            lineage_repository_max_file_count=100,
            lineage_script_max_bytes=1024,
        ),
    )
    result = lineage_jobs.script_repository_sync_handler(db_session, job)

    assert result["deleted_count"] == 1
    assert added.script_file.enabled is False
    deleted = db_session.scalar(select(ScriptChangeSet).where(
        ScriptChangeSet.script_file_id == added.script_file.id,
        ScriptChangeSet.change_type == "deleted",
    ))
    assert deleted is not None


def test_lineage_export_is_a_real_workbook_with_required_sheets(tmp_path: Path, monkeypatch) -> None:
    import app.api.lineage as lineage_api
    import app.services.lineage.jobs as lineage_jobs

    storage = LocalStorageService(tmp_path / "storage")
    monkeypatch.setattr(lineage_api, "get_storage_service", lambda: storage)
    monkeypatch.setattr(lineage_jobs, "get_storage_service", lambda: storage)
    with _client() as client:
        project = client.post("/api/projects", json={"name": "export", "institution_id": 1}).json()
        client.post(
            f"/api/projects/{project['id']}/scripts/upload",
            data={"relative_path": "load.sql", "dialect": "sqlite"},
            files={"file": ("load.sql", b"insert into MART.T (A) select A from ODS.S", "text/plain")},
        )
        response = client.get(f"/api/projects/{project['id']}/export/lineage-workbook")
        assert response.status_code == 200, response.text
        workbook = load_workbook(BytesIO(response.content))
        assert workbook.sheetnames == [
            "血缘总览", "字段级血缘", "表级血缘", "脚本清单", "脚本依赖", "加工逻辑",
            "未解析节点", "版本变更", "影响分析", "待确认问题", "审核记录",
        ]
        assert workbook["字段级血缘"].freeze_panes == "A2"
        assert workbook["字段级血缘"].auto_filter.ref
        queued = client.post(f"/api/projects/{project['id']}/export/lineage-workbook/jobs")
        assert queued.status_code == 200, queued.text
        assert queued.json()["status"] == "completed"
        assert queued.json()["result"]["file_id"]


def test_case_coalesce_and_window_expressions_keep_all_source_columns() -> None:
    result = parse_sql_lineage("""
        insert into MART.T (CODE_NAME, BEST_PHONE, RN)
        select case when s.CODE='1' then 'A' else 'B' end,
               coalesce(s.MOBILE, s.PHONE),
               row_number() over (partition by s.CUSTOMER_ID order by s.UPDATED_AT desc)
        from ODS.S s
    """, dialect="sqlite")
    by_target: dict[str, list] = {}
    for edge in result.edges:
        by_target.setdefault(edge.target.logical_name, []).append(edge)
    assert {item.source.column_name for item in by_target["MART.T.CODE_NAME"]} == {"CODE"}
    assert all(item.edge_type == "maps_code" for item in by_target["MART.T.CODE_NAME"])
    assert {item.source.column_name for item in by_target["MART.T.BEST_PHONE"]} == {"MOBILE", "PHONE"}
    assert {item.source.column_name for item in by_target["MART.T.RN"]} == {"CUSTOMER_ID", "UPDATED_AT"}


def test_union_lineage_keeps_sources_from_both_branches() -> None:
    result = parse_sql_lineage("insert into MART.T (A) select A from ODS.S1 union all select A from ODS.S2", dialect="sqlite")
    sources = {edge.source.logical_name for edge in result.edges if edge.target.logical_name == "MART.T.A"}
    assert sources == {"ODS.S1.A", "ODS.S2.A"}


def test_merge_update_and_insert_produce_column_lineage() -> None:
    result = parse_sql_lineage("""
        merge into MART.T t using ODS.S s on t.ID=s.ID
        when matched then update set t.A=s.A
        when not matched then insert (ID,A) values (s.ID,s.A)
    """, dialect="postgres")
    pairs = {(edge.source.logical_name, edge.target.logical_name) for edge in result.edges}
    assert ("ODS.S.A", "MART.T.A") in pairs
    assert ("ODS.S.ID", "MART.T.ID") in pairs
    assert all(edge.join_condition and "ID" in edge.join_condition for edge in result.edges)


def test_update_from_produces_source_to_target_column_lineage() -> None:
    result = parse_sql_lineage("update MART.T t set A=s.A from ODS.S s where t.ID=s.ID", dialect="postgres")
    assert any(edge.source.logical_name == "ODS.S.A" and edge.target.logical_name == "MART.T.A" for edge in result.edges)


def test_critical_impact_marks_mappings_stale_without_overwriting_final_content(db_session: Session) -> None:
    user = User(username="impact-owner", status="active")
    project = Project(name="impact governance")
    db_session.add_all([user, project]); db_session.flush()
    target_table = TargetTable(project_id=project.id, table_code="YBT_T", table_name="YBT")
    mart_table = MartTable(project_id=project.id, table_code="MART_T", table_name="Mart")
    scenario = ProductScenario(project_id=project.id, scenario_code="DEFAULT", scenario_name="默认")
    db_session.add_all([target_table, mart_table, scenario]); db_session.flush()
    target = TargetField(project_id=project.id, target_table_id=target_table.id, field_code="A", field_name="A")
    mart = MartField(project_id=project.id, mart_table_id=mart_table.id, field_code="A", field_name="A")
    db_session.add_all([target, mart]); db_session.flush()
    technical = ScenarioTechnicalLineage(project_id=project.id, target_field_id=target.id, scenario_id=scenario.id, final_content="人工技术口径")
    business = ScenarioBusinessMapping(project_id=project.id, target_field_id=target.id, scenario_id=scenario.id, final_content="人工业务口径")
    source_mapping = SourceToMartMapping(project_id=project.id, mart_field_id=mart.id, final_content="人工入集市口径")
    ybt_mapping = MartToYbtMapping(project_id=project.id, target_field_id=target.id, mart_field_id=mart.id, final_content="人工上报口径")
    script = ScriptFileModel(project_id=project.id, relative_path="load.sql", file_name="load.sql", file_type="sql", current_version_no=2)
    db_session.add_all([business, technical, source_mapping, ybt_mapping, script]); db_session.flush()
    old = ScriptFileVersion(project_id=project.id, script_file_id=script.id, version_no=1, file_hash="a"*64, normalized_hash="b"*64, raw_content_storage_file_id=1, parse_status="parsed", created_by=user.id)
    new = ScriptFileVersion(project_id=project.id, script_file_id=script.id, version_no=2, file_hash="c"*64, normalized_hash="d"*64, raw_content_storage_file_id=2, parse_status="parsed", created_by=user.id)
    db_session.add_all([old, new]); db_session.flush()
    mart_node = LineageNode(project_id=project.id, node_type="column", logical_name="MART_T.A", table_name="MART_T", column_name="A", mart_field_id=mart.id, script_file_id=script.id, script_file_version_id=old.id, unresolved_flag=False)
    target_node = LineageNode(project_id=project.id, node_type="column", logical_name="YBT_T.A", table_name="YBT_T", column_name="A", target_field_id=target.id, script_file_id=script.id, script_file_version_id=new.id, unresolved_flag=False)
    db_session.add_all([mart_node, target_node]); db_session.flush()
    db_session.add(LineageEdge(
        project_id=project.id, script_file_version_id=new.id,
        source_node_id=mart_node.id, target_node_id=target_node.id,
        edge_type="derives_from", confidence_level="high", enabled=True,
    ))
    db_session.flush()

    diff = compare_sql_versions("insert into MART_T(A,B) select A,B from ODS_S", "insert into MART_T(A) select A from ODS_S", dialect="sqlite")
    _change, impact = persist_change_impact(db_session, script_file=script, from_version=old, to_version=new, diff=diff, created_by=user.id)

    assert impact.severity == "critical"
    assert impact.affected_scenario_mapping_ids_json == [business.id]
    assert f"scenario_business:{business.id}" in impact.affected_mapping_ids_json
    assert technical.lineage_status == source_mapping.lineage_status == ybt_mapping.lineage_status == "stale"
    assert (technical.final_content, source_mapping.final_content, ybt_mapping.final_content) == ("人工技术口径", "人工入集市口径", "人工上报口径")
    workbook = load_workbook(BytesIO(export_lineage_workbook(db_session, project.id)))
    headers = [cell.value for cell in workbook["字段级血缘"][1]]
    status_column = headers.index("血缘状态") + 1
    assert workbook["字段级血缘"].cell(2, status_column).value == "stale"
    assert db_session.query(ReviewTask).count() == 3


def test_task_payload_redaction_preserves_safe_storage_keys_but_not_sensitive_text() -> None:
    digest = "807914e0d8a5e97b79d85f8497a6c34c6dd856577473afb1234567890abcdef"
    storage_key = f"projects/1/{digest[:2]}/{digest}.xlsx"
    result = redact_summary({"storage_key": storage_key, "note": "账号 6222020202020202", "password": "never-store"})
    assert result["storage_key"] == storage_key
    assert "6222020202020202" not in result["note"]
    assert "password" not in result


def test_shell_call_change_is_a_high_risk_dependency_change() -> None:
    result = compare_shell_versions("sh load_a.sh TABLE_A", "sh load_b.sh TABLE_B")
    assert result.semantic_changed is True
    assert [item.change_category for item in result.items] == ["script_dependency_changed"]
    assert result.severity == "high"


@contextmanager
def _client() -> Iterator[TestClient]:
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, autoflush=False)
    with factory() as seed:
        seed.add(Institution(institution_code="TEST_BANK", institution_name="测试银行"))
        seed.commit()

    def override() -> Iterator[Session]:
        session = factory()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = override
    try:
        with TestClient(app) as client:
            yield client
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(engine)
import subprocess
