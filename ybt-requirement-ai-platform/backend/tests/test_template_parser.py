from pathlib import Path

from openpyxl import Workbook

from app.services.template_parser.excel_parser import ExcelTemplateParser


def test_excel_template_parser_reads_xlsx_multi_sheet(tmp_path: Path):
    file_path = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "客户信息表"
    sheet.append(["表编号", "表名称", "字段代码", "字段中文名", "字段类型", "是否必填", "字段定义", "监管说明"])
    sheet.append(["YBT_CUSTOMER", "客户信息表", "CERT_TYPE", "客户证件类型", "varchar(20)", "是", "证件类型代码", "按监管代码集转换"])
    second = workbook.create_sheet("贷款信息表")
    second.append(["字段代码", "字段名称", "数据类型", "必填", "填报说明", "校验规则"])
    second.append(["LOAN_BAL", "贷款余额", "decimal(18,2)", "Y", "贷款余额字段", "余额不小于0"])
    workbook.save(file_path)

    output = ExcelTemplateParser().parse(str(file_path))

    assert output.sheet_count == 2
    assert output.field_count == 2
    assert output.results[0].table_code == "YBT_CUSTOMER"
    assert output.results[0].table_name == "客户信息表"
    assert output.results[0].parsed_rows[0]["field_code"] == "CERT_TYPE"
    assert output.results[1].table_name == "贷款信息表"


def test_excel_template_parser_warns_when_field_code_missing(tmp_path: Path):
    file_path = tmp_path / "bad-template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "客户信息表"
    sheet.append(["字段中文名", "字段类型"])
    sheet.append(["客户证件类型", "varchar(20)"])
    workbook.save(file_path)

    output = ExcelTemplateParser().parse(str(file_path))

    assert output.results[0].warnings
    assert "缺少字段代码" in output.results[0].warnings[0]
