from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ProductScenario, ScenarioBusinessMapping, ScenarioTechnicalLineage, TargetField, TargetTable

FIXED_HEADERS = [
    ("field_code", "数据项编码"), ("field_name", "数据项名称"), ("data_category", "数据类别"),
    ("data_format", "数据格式"), ("regulatory_original_definition", "字段业务定义（监管原始口径）"),
    ("regulatory_refined_definition", "字段业务定义（监管定义细化）"), ("report_name", "报表名称"),
    ("report_field_name", "字段名称"), ("east_definition", "EAST口径"),
    ("internal_definition", "字段业务定义（行内）"), ("remarks", "备注"),
]
BUSINESS_HEADERS = [
    ("business_definition", "字段业务定义"), ("source_system_screenshot_required", "源系统截图"),
    ("source_system_change_required", "源系统改造"), ("external_data_required", "外部数据"),
    ("manual_supplement_required", "手工补录"), ("business_owner", "业务口径确认人"), ("remarks", "备注"),
]
TECHNICAL_HEADERS = [
    ("source_system_name", "来源系统"), ("source_database_name", "来源库"), ("source_schema_name", "来源schema"),
    ("source_table_english_name", "来源表英文名"), ("source_table_chinese_name", "来源表中文名"),
    ("source_field_english_name", "来源字段英文名"), ("source_field_chinese_name", "来源字段中文名"),
    ("processing_logic", "处理逻辑"), ("processing_logic_type", "处理逻辑类型"),
    ("tech_owner", "技术口径确认人"), ("remarks", "备注"),
]


def export_traceability_workbook(db: Session, project_id: int, table_id: int | None = None) -> bytes:
    table_statement = select(TargetTable).where(TargetTable.project_id == project_id)
    if table_id is not None:
        table_statement = table_statement.where(TargetTable.id == table_id)
    tables = list(db.scalars(table_statement.order_by(TargetTable.id)).all())
    table_ids = [table.id for table in tables]
    fields = list(db.scalars(select(TargetField).where(TargetField.target_table_id.in_(table_ids)).order_by(
        TargetField.target_table_id, TargetField.id
    )).all()) if table_ids else []
    scenarios = list(db.scalars(select(ProductScenario).where(
        ProductScenario.project_id == project_id, ProductScenario.enabled.is_(True)
    ).order_by(ProductScenario.sort_order, ProductScenario.id)).all())
    field_ids = [field.id for field in fields]
    businesses = list(db.scalars(select(ScenarioBusinessMapping).where(
        ScenarioBusinessMapping.target_field_id.in_(field_ids)
    )).all()) if field_ids else []
    lineages = list(db.scalars(select(ScenarioTechnicalLineage).where(
        ScenarioTechnicalLineage.target_field_id.in_(field_ids)
    )).all()) if field_ids else []
    business_index = {(item.target_field_id, item.scenario_id): item for item in businesses}
    lineage_index = {(item.target_field_id, item.scenario_id): item for item in lineages}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "业务口径及技术溯源"
    _write_headers(sheet, scenarios)
    for row_number, field in enumerate(fields, start=3):
        column = 1
        for key, _ in FIXED_HEADERS:
            value = getattr(field, key, None)
            if key == "data_format" and not value:
                value = field.field_type
            if key == "regulatory_original_definition" and not value:
                value = field.regulatory_description
            if key == "internal_definition" and not value:
                value = field.field_definition
            sheet.cell(row_number, column, _display(value))
            column += 1
        for scenario in scenarios:
            business = business_index.get((field.id, scenario.id))
            lineage = lineage_index.get((field.id, scenario.id))
            for key, _ in BUSINESS_HEADERS:
                value = getattr(business, key, None) if business else None
                if key == "business_definition" and business and business.final_content:
                    value = business.final_content
                sheet.cell(row_number, column, _display(value))
                column += 1
            for key, _ in TECHNICAL_HEADERS:
                value = getattr(lineage, key, None) if lineage else None
                if key == "processing_logic" and lineage and lineage.final_content:
                    value = lineage.final_content
                sheet.cell(row_number, column, _display(value))
                column += 1
    _style_main_sheet(sheet, max(len(fields) + 2, 2))
    _write_review_sheet(workbook, fields, scenarios, business_index, lineage_index)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_headers(sheet, scenarios: list[ProductScenario]) -> None:
    column = 1
    for _, title in FIXED_HEADERS:
        sheet.cell(1, column, title)
        sheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
        column += 1
    for scenario in scenarios:
        business_start = column
        sheet.cell(1, business_start, f"业务口径-{scenario.scenario_name}")
        for offset, (_, title) in enumerate(BUSINESS_HEADERS):
            sheet.cell(2, business_start + offset, title)
        sheet.merge_cells(start_row=1, start_column=business_start, end_row=1, end_column=business_start + len(BUSINESS_HEADERS) - 1)
        column += len(BUSINESS_HEADERS)
        technical_start = column
        sheet.cell(1, technical_start, f"溯源-{scenario.scenario_name}")
        for offset, (_, title) in enumerate(TECHNICAL_HEADERS):
            sheet.cell(2, technical_start + offset, title)
        sheet.merge_cells(start_row=1, start_column=technical_start, end_row=1, end_column=technical_start + len(TECHNICAL_HEADERS) - 1)
        column += len(TECHNICAL_HEADERS)


def _style_main_sheet(sheet, max_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    group_fill = PatternFill("solid", fgColor="B8CCE4")
    for row in (1, 2):
        for cell in sheet[row]:
            cell.font = Font(bold=True)
            cell.fill = group_fill if row == 1 else header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=3, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "L3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(sheet.max_column)}{max_row}"
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 42
    for column in range(1, sheet.max_column + 1):
        title = str(sheet.cell(2, column).value or sheet.cell(1, column).value or "")
        width = 16
        if any(key in title for key in ["定义", "口径", "逻辑", "备注"]):
            width = 28
        elif any(key in title for key in ["英文名", "中文名", "确认人"]):
            width = 20
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_review_sheet(workbook, fields, scenarios, business_index, lineage_index) -> None:
    sheet = workbook.create_sheet("审核状态与待确认问题")
    headers = ["数据项编码", "数据项名称", "场景", "业务确认状态", "技术确认状态", "业务待确认问题", "技术待确认问题"]
    sheet.append(headers)
    for field in fields:
        for scenario in scenarios:
            business = business_index.get((field.id, scenario.id))
            lineage = lineage_index.get((field.id, scenario.id))
            if not business and not lineage:
                continue
            sheet.append([
                field.field_code, field.field_name, scenario.scenario_name,
                business.business_confirm_status if business else "未维护",
                lineage.tech_confirm_status if lineage else "未维护",
                business.open_questions if business else None, lineage.open_questions if lineage else None,
            ])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(sheet.max_row, 1)}"
    for column in range(1, 8):
        sheet.column_dimensions[get_column_letter(column)].width = 24 if column >= 4 else 18


def _display(value):
    if isinstance(value, bool):
        return "是" if value else "否"
    return value
