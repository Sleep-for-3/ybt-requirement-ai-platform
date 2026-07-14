from pathlib import Path
from uuid import uuid4
from datetime import UTC, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from app.core.settings import get_settings
from app.models import CatalogColumn, CatalogSchema, CatalogTable, MetadataImportDocument
from app.services.metadata.hashing import metadata_hash

ALIASES={"system_name":{"系统名称","系统名"},"datasource_name":{"数据源名称","数据源名"},"database_name":{"数据库名","库名","database","database_name"},"schema_name":{"schema","模式名","schema_name"},"table_name":{"表英文名","表名","table_name"},"table_comment":{"表中文名","表注释","表说明"},"column_name":{"字段英文名","字段名","column_name"},"column_comment":{"字段中文名","字段注释","字段说明"},"data_type":{"字段类型","数据类型","data_type"},"nullable":{"是否可空","可空"},"is_primary_key":{"主键","是否主键"},"ordinal_position":{"字段顺序","字段序号"}}

async def ingest_metadata_excel(db,datasource,upload):
    if Path(upload.filename or "").suffix.lower() != ".xlsx": raise ValueError("元数据字典只支持 .xlsx")
    content=await upload.read(); directory=Path(get_settings().storage_dir)/"projects"/str(datasource.project_id)/"metadata-imports"; directory.mkdir(parents=True,exist_ok=True)
    path=directory/f"{uuid4().hex}-{(upload.filename or 'metadata.xlsx').replace('/','_').replace(chr(92),'_')}"; path.write_bytes(content)
    rows,warnings=_parse(path); doc=MetadataImportDocument(project_id=datasource.project_id,datasource_id=datasource.id,file_name=upload.filename or "metadata.xlsx",storage_path=str(path),parse_status="success",parse_summary_json={"row_count":len(rows),"sheet_count":len({x['source_sheet'] for x in rows})},parsed_rows_json=rows,warnings_json=warnings)
    db.add(doc);db.commit();db.refresh(doc);return doc

def _parse(path):
    workbook=load_workbook(path,data_only=True); output=[];warnings=[]
    normalized={alias.lower().replace(" ",""):field for field,aliases in ALIASES.items() for alias in aliases}
    for sheet in workbook.worksheets:
        header_row=None;mapping={}
        for row in range(1,min(sheet.max_row,30)+1):
            candidate={column:normalized.get(str(sheet.cell(row,column).value or "").strip().lower().replace(" ","")) for column in range(1,sheet.max_column+1)}
            if "table_name" in candidate.values() and "column_name" in candidate.values(): header_row=row;mapping={column:field for column,field in candidate.items() if field};break
        if header_row is None:warnings.append(f"{sheet.title} 未识别表头");continue
        for row in range(header_row+1,sheet.max_row+1):
            item={field:sheet.cell(row,column).value for column,field in mapping.items()}
            if not item.get("table_name") or not item.get("column_name"):continue
            item["schema_name"]=str(item.get("schema_name") or "main");item["source_sheet"]=sheet.title;item["source_row"]=row;item["source_cells"]={field:f"{sheet.title}!{get_column_letter(column)}{row}" for column,field in mapping.items()};output.append(item)
    return output,warnings

def apply_metadata_excel(db,document):
    now=datetime.now(UTC);schemas=tables=columns=0;seen_schemas=set();seen_tables=set()
    for row in document.parsed_rows_json or []:
        schema_name=str(row.get("schema_name") or "main");table_name=str(row["table_name"]);column_name=str(row["column_name"])
        schema=db.scalar(select(CatalogSchema).where(CatalogSchema.datasource_id==document.datasource_id,CatalogSchema.schema_name==schema_name))
        if schema is None:schema=CatalogSchema(project_id=document.project_id,datasource_id=document.datasource_id,schema_name=schema_name);db.add(schema);db.flush();schemas+=1
        table=db.scalar(select(CatalogTable).where(CatalogTable.datasource_id==document.datasource_id,CatalogTable.schema_name==schema_name,CatalogTable.table_name==table_name))
        if table is None:table=CatalogTable(project_id=document.project_id,datasource_id=document.datasource_id,catalog_schema_id=schema.id,database_name=row.get("database_name"),schema_name=schema_name,table_name=table_name);db.add(table);db.flush();tables+=1
        table.database_name=row.get("database_name") or table.database_name
        table.table_comment=row.get("table_comment");table.enabled=True;table.last_synced_at=now;table.metadata_hash=metadata_hash({"schema":schema_name,"table":table_name,"comment":table.table_comment})
        column=db.scalar(select(CatalogColumn).where(CatalogColumn.catalog_table_id==table.id,CatalogColumn.column_name==column_name))
        if column is None:column=CatalogColumn(project_id=document.project_id,datasource_id=document.datasource_id,catalog_table_id=table.id,database_name=table.database_name,schema_name=schema_name,table_name=table_name,column_name=column_name);db.add(column);columns+=1
        column.database_name=table.database_name
        column.column_comment=row.get("column_comment");column.data_type=str(row.get("data_type") or "") or None;column.database_native_type=column.data_type;column.nullable=_bool(row.get("nullable"),True);column.is_primary_key=_bool(row.get("is_primary_key"),False);column.ordinal_position=int(row.get("ordinal_position") or 0);column.enabled=True;column.last_synced_at=now;column.metadata_hash=metadata_hash(row)
    db.commit();return {"document_id":document.id,"schemas":schemas,"tables":tables,"columns":columns,"warnings":document.warnings_json or []}

def _bool(value,default):
    if value is None or value=="":return default
    return str(value).strip().lower() in {"是","y","yes","true","1"}
