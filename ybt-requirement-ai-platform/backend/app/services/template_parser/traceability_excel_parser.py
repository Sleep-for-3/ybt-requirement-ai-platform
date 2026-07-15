from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FIXED_COLUMNS = {
    "field_code": {"数据项编码"}, "field_name": {"数据项名称"}, "data_category": {"数据类别"},
    "data_format": {"数据格式"},
    "regulatory_original_definition": {"字段业务定义（监管原始口径）", "字段业务定义(监管原始口径)"},
    "regulatory_refined_definition": {"字段业务定义（监管定义细化）", "字段业务定义(监管定义细化)"},
    "report_name": {"报表名称"}, "report_field_name": {"字段名称"},
    "east_definition": {"EAST口径", "EAST 口径"},
    "internal_definition": {"字段业务定义（行内）", "字段业务定义(行内)"}, "remarks": {"备注"},
}
BUSINESS_COLUMNS = {
    "字段业务定义": "business_definition", "源系统截图": "source_system_screenshot_required",
    "源系统改造": "source_system_change_required", "外部数据": "external_data_required",
    "手工补录": "manual_supplement_required", "业务口径确认人": "business_owner",
    "备注": "remarks", "待确认问题": "open_questions",
}
TECHNICAL_COLUMNS = {
    "来源系统": "source_system_name", "来源库": "source_database_name", "来源schema": "source_schema_name",
    "来源Schema": "source_schema_name", "来源表英文名": "source_table_english_name",
    "来源表中文名": "source_table_chinese_name", "来源字段英文名": "source_field_english_name",
    "来源字段中文名": "source_field_chinese_name", "处理逻辑": "processing_logic",
    "处理逻辑类型": "processing_logic_type", "技术口径确认人": "tech_owner",
    "备注": "remarks", "待确认问题": "open_questions",
}
SCENARIO_CODES = {
    "借记卡": "DEBIT_CARD", "信用卡": "CREDIT_CARD", "储蓄存款": "SAVING_DEPOSIT",
    "贷款产品": "LOAN_PRODUCT", "代销保险": "AGENT_INSURANCE", "代销理财": "AGENT_WEALTH",
    "代销基金": "AGENT_FUND", "债券回购": "BOND_REPO", "同业拆借": "INTERBANK_LENDING",
    "发行理财": "ISSUED_WEALTH", "手工补录": "MANUAL_SUPPLEMENT", "其他场景": "OTHER", "其他": "OTHER",
}


@dataclass
class TraceabilitySheetResult:
    sheet_name: str
    header_start_row: int
    header_end_row: int
    fixed_columns: list[dict[str, Any]]
    scenario_groups: list[dict[str, Any]]
    parsed_rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)


@dataclass
class TraceabilityParseOutput:
    sheet_count: int
    row_count: int
    sheet_names: list[str]
    detected_scenarios: list[dict[str, str]]
    results: list[TraceabilitySheetResult]
    warnings: list[str]


class TraceabilityExcelParser:
    def parse(self, file_path: Any) -> TraceabilityParseOutput:
        source = file_path
        if isinstance(file_path, (str, Path)):
            path = Path(file_path)
            if path.suffix.lower() != ".xlsx":
                raise ValueError("业务口径及溯源表只支持 .xlsx 文件。")
            source = path
        workbook = load_workbook(source, data_only=True)
        results = []
        for sheet in workbook.worksheets:
            try:
                results.append(self._parse_sheet(sheet))
            except Exception as exc:
                results.append(TraceabilitySheetResult(
                    sheet_name=sheet.title,
                    header_start_row=1,
                    header_end_row=1,
                    fixed_columns=[],
                    scenario_groups=[],
                    parsed_rows=[],
                    warnings=[f"Sheet {sheet.title} 解析失败，已跳过：{exc}"],
                ))
        scenarios: dict[str, dict[str, str]] = {}
        for result in results:
            for group in result.scenario_groups:
                scenarios[group["scenario_code"]] = {
                    "scenario_code": group["scenario_code"], "scenario_name": group["scenario_name"]
                }
        return TraceabilityParseOutput(
            sheet_count=len(results), row_count=sum(len(item.parsed_rows) for item in results),
            sheet_names=[sheet.title for sheet in workbook.worksheets], detected_scenarios=list(scenarios.values()),
            results=results, warnings=[warning for item in results for warning in item.warnings],
        )

    def _parse_sheet(self, sheet) -> TraceabilitySheetResult:
        matrix = _expanded_matrix(sheet)
        header_start = _find_header_start(matrix)
        if header_start is None:
            warning = f"Sheet {sheet.title} 未识别到数据项编码或场景分组表头"
            return TraceabilitySheetResult(sheet.title, 1, 1, [], [], [], [warning])
        header_end = _find_header_end(matrix, header_start)
        fixed_map = _fixed_column_map(matrix, header_start, header_end)
        group_columns, warnings = _scenario_column_map(matrix, header_start, header_end)
        fixed_columns = [
            {"field": name, "column": column, "header": _cell_text(matrix[header_start - 1][column - 1])}
            for name, column in fixed_map.items()
        ]
        if "field_code" not in fixed_map:
            warnings.append(f"Sheet {sheet.title} 缺少关键列：数据项编码")
        if "field_name" not in fixed_map:
            warnings.append(f"Sheet {sheet.title} 缺少关键列：数据项名称")
        parsed_rows: list[dict[str, Any]] = []
        for row_number in range(header_end + 1, len(matrix) + 1):
            values = matrix[row_number - 1]
            if not any(_cell_text(value) for value in values):
                continue
            fixed, sources = {}, {}
            for name, column in fixed_map.items():
                fixed[name] = _cell_text(values[column - 1])
                sources[f"fixed.{name}"] = _location(sheet.title, row_number, column)
            scenarios: dict[str, dict[str, Any]] = {}
            for info in group_columns:
                scenario = scenarios.setdefault(info["scenario_name"], {
                    "scenario_code": info["scenario_code"], "business": {}, "technical": {}
                })
                value = _cell_text(values[info["column"] - 1])
                if value:
                    scenario[info["layer"]][info["field"]] = _yes_no(value) if info["field"].endswith("_required") else value
                    sources[f"scenarios.{info['scenario_name']}.{info['layer']}.{info['field']}"] = _location(
                        sheet.title, row_number, info["column"]
                    )
            if fixed.get("field_code") or fixed.get("field_name") or any(item["business"] or item["technical"] for item in scenarios.values()):
                if not fixed.get("field_code"):
                    warnings.append(f"Sheet {sheet.title} 第 {row_number} 行缺少数据项编码，apply 时将跳过")
                parsed_rows.append({"row_number": row_number, "fixed": fixed, "scenarios": scenarios, "sources": sources})
        return TraceabilitySheetResult(
            sheet.title, header_start, header_end, fixed_columns, _scenario_group_summary(group_columns), parsed_rows, warnings
        )


def _expanded_matrix(sheet) -> list[list[Any]]:
    matrix = [[sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)] for row in range(1, sheet.max_row + 1)]
    for merged in sheet.merged_cells.ranges:
        value = matrix[merged.min_row - 1][merged.min_col - 1]
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                matrix[row - 1][column - 1] = value
    return matrix


def _find_header_start(matrix: list[list[Any]]) -> int | None:
    aliases = {_normal(alias) for values in FIXED_COLUMNS.values() for alias in values}
    for number, row in enumerate(matrix, start=1):
        if {_normal(value) for value in row} & aliases or any(_group_title(value) for value in row):
            return number
    return None


def _find_header_end(matrix: list[list[Any]], start: int) -> int:
    if not any(_group_title(value) for value in matrix[start - 1]):
        return start
    known = set(BUSINESS_COLUMNS) | set(TECHNICAL_COLUMNS)
    for number in range(start + 1, min(start + 4, len(matrix)) + 1):
        if any(_cell_text(value) in known for value in matrix[number - 1]):
            return number
    return start


def _fixed_column_map(matrix: list[list[Any]], start: int, end: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(1, len(matrix[0]) + 1):
        headers = {_normal(matrix[row - 1][column - 1]) for row in range(start, end + 1)}
        for name, aliases in FIXED_COLUMNS.items():
            if name not in result and headers & {_normal(alias) for alias in aliases}:
                result[name] = column
    return result


def _scenario_column_map(matrix: list[list[Any]], start: int, end: int) -> tuple[list[dict[str, Any]], list[str]]:
    columns, warnings = [], []
    group_occurrences: dict[tuple[str, str], set[str]] = {}
    for column in range(1, len(matrix[0]) + 1):
        headers = [_cell_text(matrix[row - 1][column - 1]) for row in range(start, end + 1)]
        title = next((value for value in headers if _group_title(value)), None)
        if title is None:
            continue
        layer, raw_name = _group_title(title) or ("", "")
        scenario_name = _scenario_name(raw_name)
        mapping = BUSINESS_COLUMNS if layer == "business" else TECHNICAL_COLUMNS
        subheader = next((value for value in reversed(headers) if value in mapping), "")
        if not subheader:
            warnings.append(f"列 {get_column_letter(column)} 的场景子表头未识别，已跳过")
            continue
        key = (scenario_name, layer)
        seen_titles = group_occurrences.setdefault(key, set())
        if seen_titles and title not in seen_titles:
            warnings.append(f"检测到重复场景分组：{scenario_name} / {layer}")
        seen_titles.add(title)
        columns.append({"column": column, "layer": layer, "field": mapping[subheader], "scenario_name": scenario_name,
                        "scenario_code": _scenario_code(scenario_name), "group_title": title, "subheader": subheader})
    return columns, warnings


def _scenario_group_summary(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for item in columns:
        group = groups.setdefault(item["scenario_name"], {
            "scenario_code": item["scenario_code"], "scenario_name": item["scenario_name"],
            "has_business": False, "has_technical": False, "business_columns": [], "technical_columns": [],
        })
        group[f"has_{item['layer']}"] = True
        group[f"{item['layer']}_columns"].append({"field": item["field"], "column": item["column"]})
    return list(groups.values())


def _group_title(value: Any) -> tuple[str, str] | None:
    match = re.match(r"^(业务口径|溯源)\s*[-－—:：]\s*(.+)$", _cell_text(value))
    if not match:
        return None
    return ("business" if match.group(1) == "业务口径" else "technical", match.group(2).strip())


def _scenario_name(value: str) -> str:
    name = value.strip()
    if name.endswith("系统"):
        name = name[:-2]
    return {"贷记卡": "信用卡", "信贷": "贷款产品"}.get(name, name)


def _scenario_code(name: str) -> str:
    if name in SCENARIO_CODES:
        return SCENARIO_CODES[name]
    ascii_code = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").upper()
    return ascii_code or f"SCENARIO_{sum(ord(char) for char in name):X}"


def _normal(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_text(value)).replace("（", "(").replace("）", ")").lower()


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _yes_no(value: str) -> bool:
    return value.strip().lower() in {"是", "y", "yes", "true", "1", "需要"}


def _location(sheet_name: str, row: int, column: int) -> str:
    return f"{sheet_name}!{get_column_letter(column)}{row}"
