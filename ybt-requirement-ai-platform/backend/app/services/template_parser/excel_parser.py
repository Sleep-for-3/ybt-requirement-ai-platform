from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COLUMN_ALIASES = {
    "table_code": {"表编号", "表代码", "数据表编号", "监管表编号", "table_code"},
    "table_name": {"表名称", "表名", "数据表名称", "监管表名称", "table_name"},
    "field_code": {"字段编号", "字段代码", "字段名", "字段英文名", "字段标识", "field_code"},
    "field_name": {"字段名称", "字段中文名", "中文名称", "字段说明", "field_name"},
    "field_type": {"字段类型", "数据类型", "类型", "field_type"},
    "required_flag": {"是否必填", "必填", "是否为空", "是否可空", "required", "required_flag"},
    "field_definition": {"字段定义", "业务定义", "口径说明", "填报说明", "采集口径", "field_definition"},
    "regulatory_description": {"监管说明", "监管口径", "报送说明", "校验规则", "校验规则说明", "regulatory_description"},
}


@dataclass
class ExcelSheetParseResult:
    sheet_name: str
    table_code: str | None
    table_name: str | None
    field_count: int
    raw_header: list[str]
    parsed_rows: list[dict[str, Any]]
    warnings: list[str] = field(default_factory=list)


@dataclass
class ExcelTemplateParseOutput:
    sheet_count: int
    table_count: int
    field_count: int
    sheet_names: list[str]
    results: list[ExcelSheetParseResult]
    warnings: list[str]


class ExcelTemplateParser:
    def parse(self, file_path: str) -> ExcelTemplateParseOutput:
        path = Path(file_path)
        if path.suffix.lower() == ".xls":
            raise ValueError("暂不支持 .xls，请另存为 .xlsx 后上传。")
        if path.suffix.lower() != ".xlsx":
            raise ValueError("MVP 阶段只支持 .xlsx 一表通模板。")

        workbook = load_workbook(path, data_only=True)
        results: list[ExcelSheetParseResult] = []
        all_warnings: list[str] = []
        for sheet in workbook.worksheets:
            result = self._parse_sheet(sheet)
            results.append(result)
            all_warnings.extend(result.warnings)
        return ExcelTemplateParseOutput(
            sheet_count=len(results),
            table_count=sum(1 for result in results if result.field_count > 0),
            field_count=sum(result.field_count for result in results),
            sheet_names=[sheet.title for sheet in workbook.worksheets],
            results=results,
            warnings=all_warnings,
        )

    def _parse_sheet(self, sheet) -> ExcelSheetParseResult:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = _find_header_index(rows)
        if header_index is None:
            warning = f"Sheet {sheet.title} 未识别到表头"
            return ExcelSheetParseResult(sheet.title, None, sheet.title, 0, [], [], [warning])

        raw_header = [_cell_to_text(value) for value in rows[header_index]]
        column_map = _build_column_map(raw_header)
        parsed_rows: list[dict[str, Any]] = []
        warnings: list[str] = []
        for excel_row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(_cell_to_text(value) for value in values):
                continue
            parsed = _parse_row(values, column_map)
            if not parsed.get("field_code"):
                warnings.append(f"Sheet {sheet.title} 第 {excel_row_number} 行缺少字段代码，apply 时将跳过")
            if not parsed.get("field_name"):
                warnings.append(f"Sheet {sheet.title} 第 {excel_row_number} 行缺少字段名称，apply 时将跳过")
            parsed["row_number"] = excel_row_number
            parsed_rows.append(parsed)

        first_with_table = next((row for row in parsed_rows if row.get("table_code") or row.get("table_name")), {})
        table_code = first_with_table.get("table_code") or _guess_table_code(sheet.title)
        table_name = first_with_table.get("table_name") or sheet.title
        valid_field_count = sum(1 for row in parsed_rows if row.get("field_code") and row.get("field_name"))
        return ExcelSheetParseResult(
            sheet_name=sheet.title,
            table_code=table_code,
            table_name=table_name,
            field_count=valid_field_count,
            raw_header=raw_header,
            parsed_rows=parsed_rows,
            warnings=warnings,
        )


def _find_header_index(rows: list[tuple]) -> int | None:
    alias_values = {alias for aliases in COLUMN_ALIASES.values() for alias in aliases}
    for index, row in enumerate(rows[:20]):
        headers = {_cell_to_text(value) for value in row}
        if headers & alias_values:
            return index
    return None


def _build_column_map(raw_header: list[str]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, name in enumerate(raw_header):
        normalized = name.strip()
        for canonical, aliases in COLUMN_ALIASES.items():
            if normalized in aliases and canonical not in column_map:
                column_map[canonical] = index
    return column_map


def _parse_row(values: tuple, column_map: dict[str, int]) -> dict[str, Any]:
    parsed: dict[str, Any] = {}
    for key in COLUMN_ALIASES:
        index = column_map.get(key)
        parsed[key] = _cell_to_text(values[index]) if index is not None and index < len(values) else ""
    parsed["required_flag"] = _parse_required_flag(parsed.get("required_flag", ""))
    return parsed


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_required_flag(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"是", "y", "yes", "true", "1", "必填", "not null", "非空"}:
        return True
    if normalized in {"否", "n", "no", "false", "0", "可空", "nullable"}:
        return False
    if "不可空" in normalized or "不能为空" in normalized:
        return True
    return False


def _guess_table_code(sheet_name: str) -> str | None:
    cleaned = sheet_name.strip()
    return cleaned if cleaned.isascii() and cleaned else None
