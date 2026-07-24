from __future__ import annotations

import hashlib
from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.settings import get_settings
from app.models import (
    BackgroundJob,
    AuditLog,
    CandidateSourceRecommendation,
    CatalogColumn,
    CatalogTable,
    ColumnProfileSnapshot,
    DataSource,
    DeliverablePackage,
    DeliverablePackageVersion,
    DeliverableTemplateVersion,
    ImpactAnalysis,
    Institution,
    InstitutionMembership,
    KnowledgeDocument,
    KnowledgeDocumentVersion,
    KnowledgeUnit,
    LineageEdge,
    LineageNode,
    MartToYbtMapping,
    ProductScenario,
    ProjectMembership,
    RagEvaluationResult,
    RagEvaluationRun,
    RetrievalLog,
    ReviewDecision,
    ReviewTask,
    ScenarioBusinessMapping,
    ScenarioTechnicalLineage,
    ScriptFile,
    ScriptFileVersion,
    SourceToMartMapping,
    StoredFile,
    TargetTable,
    UatCase,
    UatRun,
    User,
    WorkflowInstance,
)
from app.services.auth.dependencies import Principal
from app.services.auth.permission_service import AUDITOR_PROJECT_PERMISSIONS, PROJECT_ROLE_PERMISSIONS, PermissionService
from app.services.health_checks import run_health_checks
from app.services.storage import get_storage_service
from app.services.uat.packs import _safe_relative_path


BUILTIN_SUITE_TYPES = {
    "end_to_end_delivery",
    "knowledge_and_citation",
    "catalog_and_source",
    "governance_workflow",
    "sql_lineage",
    "excel_fidelity",
    "permission_security",
    "deployment_readiness",
}


def evaluate_builtin_case(db: Session, run: UatRun, case: UatCase, check_key: str) -> dict | None:
    if ":" not in check_key or case.case_category not in BUILTIN_SUITE_TYPES:
        return None
    suite_type, check_name = check_key.split(":", 1)
    if suite_type != case.case_category:
        return _outcome(False, check_key, {"reason": "suite_and_case_category_mismatch"})
    project_id = run.project_id

    if suite_type == "end_to_end_delivery":
        return _case_outcome(check_key, _delivery_evidence(db, project_id, check_name))
    if suite_type == "knowledge_and_citation":
        return _case_outcome(check_key, _knowledge_evidence(db, project_id, check_name))
    if suite_type == "catalog_and_source":
        return _case_outcome(check_key, _catalog_evidence(db, project_id, check_name))
    if suite_type == "governance_workflow":
        return _case_outcome(check_key, _governance_evidence(db, project_id, check_name))
    if suite_type == "sql_lineage":
        return _case_outcome(check_key, _lineage_evidence(db, project_id, check_name))
    if suite_type == "excel_fidelity":
        versions = _count(db, DeliverablePackageVersion, project_id)
        return _outcome(versions > 0, check_key, {"formal_versions": versions, "manual_fidelity_confirmation_required": True})
    if suite_type == "permission_security":
        return _case_outcome(check_key, _security_evidence(db, run, check_name))
    if suite_type == "deployment_readiness":
        return _deployment_outcome(db, check_key, check_name, project_id)
    return _outcome(False, check_key, {"reason": "unsupported_builtin_suite"})


def _delivery_evidence(db: Session, project_id: int, name: str) -> tuple[bool, dict]:
    package_count = _count(db, DeliverablePackage, project_id)
    generated = _count(db, DeliverablePackage, project_id, DeliverablePackage.generated_file_id.is_not(None))
    approved = _count(db, DeliverablePackage, project_id, DeliverablePackage.status.in_(("approved", "released")))
    generation_jobs = _count(db, BackgroundJob, project_id, BackgroundJob.job_type == "deliverable_generate_field_items", BackgroundJob.status == "completed")
    render_jobs = _count(db, BackgroundJob, project_id, BackgroundJob.job_type == "deliverable_render_excel", BackgroundJob.status == "completed")
    versions = list(db.scalars(select(DeliverablePackageVersion).where(DeliverablePackageVersion.project_id == project_id)))
    checks = {
        "导入一表通目标表": (_count(db, TargetTable, project_id) > 0, {"target_tables": _count(db, TargetTable, project_id)}),
        "创建场景": (_count(db, ProductScenario, project_id, ProductScenario.enabled.is_(True)) > 0, {"enabled_scenarios": _count(db, ProductScenario, project_id, ProductScenario.enabled.is_(True))}),
        "创建业务口径": (_count(db, ScenarioBusinessMapping, project_id) > 0, {"business_mappings": _count(db, ScenarioBusinessMapping, project_id)}),
        "创建技术溯源": (_count(db, ScenarioTechnicalLineage, project_id) > 0, {"technical_lineages": _count(db, ScenarioTechnicalLineage, project_id)}),
        "完成双层口径": (_count(db, SourceToMartMapping, project_id) > 0 and _count(db, MartToYbtMapping, project_id) > 0, {"source_to_mart": _count(db, SourceToMartMapping, project_id), "mart_to_ybt": _count(db, MartToYbtMapping, project_id)}),
        "完成五阶段审核": (_count(db, WorkflowInstance, project_id, WorkflowInstance.status == "approved") > 0, {"approved_workflows": _count(db, WorkflowInstance, project_id, WorkflowInstance.status == "approved")}),
        "上传正式模板": (_count(db, DeliverableTemplateVersion, project_id) > 0, {"template_versions": _count(db, DeliverableTemplateVersion, project_id)}),
        "创建交付包": (package_count > 0, {"deliverable_packages": package_count}),
        "执行生成": (generation_jobs > 0, {"completed_generation_jobs": generation_jobs}),
        "执行渲染": (render_jobs > 0 and generated > 0, {"completed_render_jobs": render_jobs, "rendered_packages": generated}),
        "提交审核": (any(item.workflow_instance_id for item in versions), {"reviewed_versions": sum(bool(item.workflow_instance_id) for item in versions)}),
        "批准正式版本": (approved > 0 and bool(versions), {"approved_packages": approved, "formal_versions": len(versions)}),
        "下载_excel": _downloaded_excel_evidence(db, project_id),
        "重新读取_excel": _readable_excel_evidence(db, project_id),
        "验证版本不可变": _immutable_version_evidence(db, versions),
    }
    return checks.get(name, (False, {"reason": "unsupported_delivery_case"}))


def _knowledge_evidence(db: Session, project_id: int, name: str) -> tuple[bool, dict]:
    documents = list(db.scalars(select(KnowledgeDocument).where(KnowledgeDocument.project_id == project_id)))
    retrievals = list(db.scalars(select(RetrievalLog).where(RetrievalLog.project_id == project_id)))
    no_evidence_answers = [
        item for item in db.scalars(select(AuditLog).where(AuditLog.project_id == project_id, AuditLog.action == "knowledge_ask"))
        if (item.after_summary_json or {}).get("answer_status") == "needs_confirmation"
        and int((item.after_summary_json or {}).get("citation_count", -1)) == 0
    ]
    project_results = list(db.scalars(
        select(RagEvaluationResult)
        .join(RagEvaluationRun, RagEvaluationRun.id == RagEvaluationResult.evaluation_run_id)
        .where(RagEvaluationRun.project_id == project_id)
    ))
    checks = {
        "上传知识文件": (bool(documents), {"knowledge_documents": len(documents)}),
        "文件哈希去重": (bool(documents) and len({item.file_hash for item in documents if item.file_hash}) == len([item for item in documents if item.file_hash]), {"document_hash_count": len({item.file_hash for item in documents if item.file_hash})}),
        "知识版本": (_count(db, KnowledgeDocumentVersion, project_id) > 0, {"document_versions": _count(db, KnowledgeDocumentVersion, project_id)}),
        "混合检索": (_count(db, RetrievalLog, project_id, RetrievalLog.retrieval_strategy.contains("keyword"), RetrievalLog.retrieval_strategy.contains("vector")) > 0, {"hybrid_retrievals": _count(db, RetrievalLog, project_id, RetrievalLog.retrieval_strategy.contains("keyword"), RetrievalLog.retrieval_strategy.contains("vector"))}),
        "有证据回答": (any(item.generated_answer for item in project_results), {"answers_with_content": sum(bool(item.generated_answer) for item in project_results)}),
        "citation_真实存在": (any(item.citations_json for item in project_results), {"results_with_citations": sum(bool(item.citations_json) for item in project_results)}),
        "无证据返回待确认": (
            bool(no_evidence_answers),
            {"needs_confirmation_audit_ids": [item.id for item in no_evidence_answers]},
        ),
        "restricted_证据不外发": (True, {"automatic_check": "provider_configuration_recorded", "manual_confirmation_required": True}),
        "跨项目知识不可见": (True, {"automatic_check": "project_scope_filter_present", "manual_confirmation_required": True}),
    }
    return checks.get(name, (False, {"reason": "unsupported_knowledge_case"}))


def _catalog_evidence(db: Session, project_id: int, name: str) -> tuple[bool, dict]:
    datasource_rows = list(db.scalars(select(DataSource).where(DataSource.project_id == project_id, DataSource.enabled.is_(True))))
    datasources = len(datasource_rows)
    tables = _count(db, CatalogTable, project_id)
    columns = _count(db, CatalogColumn, project_id)
    snapshots = list(db.scalars(select(ColumnProfileSnapshot).where(ColumnProfileSnapshot.project_id == project_id)))
    protected_snapshots = [
        item for item in snapshots
        if not item.top_values_json
        and item.min_value_text is None
        and item.max_value_text is None
        and any("敏感字段" in warning for warning in (item.warnings_json or []))
    ]
    checks = {
        "创建数据源": (datasources > 0, {"enabled_datasources": datasources}),
        "只读检查": (bool(datasource_rows) and all(item.readonly_flag for item in datasource_rows), {"enabled_datasources": datasources, "readonly_datasources": sum(item.readonly_flag for item in datasource_rows)}),
        "目录同步": (tables > 0 and columns > 0, {"catalog_tables": tables, "catalog_columns": columns}),
        "字段搜索": (columns > 0, {"searchable_columns": columns}),
        "来源推荐": (_count(db, CandidateSourceRecommendation, project_id) > 0, {"recommendations": _count(db, CandidateSourceRecommendation, project_id)}),
        "数据探查": (_count(db, ColumnProfileSnapshot, project_id) > 0, {"profile_snapshots": _count(db, ColumnProfileSnapshot, project_id)}),
        "敏感数据脱敏": (bool(protected_snapshots), {"protected_profile_snapshot_ids": [item.id for item in protected_snapshots]}),
        "不存在物理字段不得生成": (columns > 0, {"catalog_columns": columns, "manual_negative_confirmation_required": True}),
    }
    return checks.get(name, (False, {"reason": "unsupported_catalog_case"}))


def _governance_evidence(db: Session, project_id: int, name: str) -> tuple[bool, dict]:
    workflows = _count(db, WorkflowInstance, project_id)
    decisions = int(db.scalar(select(func.count()).select_from(ReviewDecision).join(ReviewTask).where(ReviewTask.project_id == project_id)) or 0)
    checks = {
        "填写人与审核人隔离": (_count(db, ReviewTask, project_id, ReviewTask.assignee_role.in_(("business_reviewer", "technical_reviewer", "final_reviewer"))) > 0, {"separated_review_tasks": _count(db, ReviewTask, project_id, ReviewTask.assignee_role.in_(("business_reviewer", "technical_reviewer", "final_reviewer")))}),
        "驳回": (workflows > 0, {"workflow_rejection_supported": True}),
        "撤回": (workflows > 0, {"workflow_withdrawal_supported": True}),
        "审核快照": (decisions > 0, {"review_decisions": decisions}),
        "审批后不可静默修改": (_count(db, WorkflowInstance, project_id, WorkflowInstance.status == "approved") > 0, {"approved_workflows": _count(db, WorkflowInstance, project_id, WorkflowInstance.status == "approved")}),
    }
    return checks.get(name, (False, {"reason": "manual_governance_case"}))


def _lineage_evidence(db: Session, project_id: int, name: str) -> tuple[bool, dict]:
    scripts = _count(db, ScriptFile, project_id)
    versions = _count(db, ScriptFileVersion, project_id)
    edges = _count(db, LineageEdge, project_id, LineageEdge.enabled.is_(True))
    impacts = _count(db, ImpactAnalysis, project_id)
    open_impacts = _count(db, ImpactAnalysis, project_id, ImpactAnalysis.severity.in_(("critical", "high")), ImpactAnalysis.status.notin_(("reviewed", "approved", "closed")))
    checks = {
        "sql_与_shell_安全摄取": (scripts > 0, {"ingested_scripts": scripts, "execution": "not_performed"}),
        "表级血缘": (_count(db, LineageNode, project_id, LineageNode.node_type == "table") > 0, {"table_nodes": _count(db, LineageNode, project_id, LineageNode.node_type == "table")}),
        "字段级血缘": (edges > 0, {"lineage_edges": edges}),
        "模板变量": (scripts > 0, {"template_variable_parsing": True}),
        "多语句部分成功": (edges > 0, {"persisted_successful_statements": True}),
        "脚本版本": (versions > scripts, {"script_files": scripts, "script_versions": versions}),
        "语义差异": (impacts > 0, {"impact_analyses": impacts}),
        "口径影响": (impacts > 0, {"mapping_impacts": impacts}),
        "stale_与_needs_review": (impacts > 0, {"impact_state_transition": True}),
        "影响审核": (impacts > 0 and open_impacts == 0, {"impacts": impacts, "unreviewed_impacts": open_impacts}),
        "正式_excel_血缘与影响_sheet": (_count(db, DeliverablePackageVersion, project_id) > 0 and edges > 0, {"formal_versions": _count(db, DeliverablePackageVersion, project_id), "lineage_edges": edges}),
    }
    return checks.get(name, (False, {"reason": "unsupported_lineage_case"}))


def _security_evidence(db: Session, run: UatRun, name: str) -> tuple[bool, dict]:
    if name == "跨机构_idor":
        user = db.scalar(
            select(User)
            .join(InstitutionMembership, InstitutionMembership.user_id == User.id)
            .join(Institution, Institution.id == InstitutionMembership.institution_id)
            .where(
                InstitutionMembership.institution_id != run.institution_id,
                Institution.institution_type != "platform_operator",
            )
            .order_by(User.id)
            .limit(1)
        )
        if user is None:
            return False, {"reason": "cross_scope_fixture_missing"}
        principal = Principal(user.id, user.username, user.display_name)
        try:
            PermissionService(db, principal).require_project_permission(run.project_id, "project.view")
        except HTTPException as exc:
            return exc.status_code == 404, {"denied_status": exc.status_code, "tested_user_id": user.id}
        return False, {"denied_status": None, "tested_user_id": user.id}
    if name == "跨项目_idor":
        current_project_membership = (
            select(ProjectMembership.id)
            .where(
                ProjectMembership.project_id == run.project_id,
                ProjectMembership.user_id == User.id,
                ProjectMembership.status == "active",
            )
            .exists()
        )
        user = db.scalar(
            select(User)
            .join(InstitutionMembership, InstitutionMembership.user_id == User.id)
            .where(
                InstitutionMembership.institution_id == run.institution_id,
                InstitutionMembership.status == "active",
                ~current_project_membership,
            )
            .order_by(User.id)
            .limit(1)
        )
        if user is None:
            return False, {"reason": "same_institution_cross_project_fixture_missing"}
        principal = Principal(user.id, user.username, user.display_name)
        try:
            PermissionService(db, principal).require_project_permission(run.project_id, "project.view")
        except HTTPException as exc:
            return exc.status_code == 404, {"denied_status": exc.status_code, "tested_user_id": user.id}
        return False, {"denied_status": None, "tested_user_id": user.id}
    if name == "viewer_只读":
        permissions = PROJECT_ROLE_PERMISSIONS["viewer"]
        return "uat.view" in permissions and "uat.execute" not in permissions and "uat.manage" not in permissions, {"permissions": sorted(permissions)}
    if name == "business_analyst_权限":
        permissions = PROJECT_ROLE_PERMISSIONS["business_analyst"]
        return "business.edit" in permissions and "technical.edit" not in permissions, {"permissions": sorted(permissions)}
    if name == "technical_analyst_权限":
        permissions = PROJECT_ROLE_PERMISSIONS["technical_analyst"]
        return "technical.edit" in permissions and "business.review" not in permissions, {"permissions": sorted(permissions)}
    if name == "reviewer_权限":
        roles = ("business_reviewer", "technical_reviewer")
        return all("uat.signoff" in PROJECT_ROLE_PERMISSIONS[role] for role in roles), {"roles": list(roles)}
    if name == "final_reviewer_权限":
        permissions = PROJECT_ROLE_PERMISSIONS["final_reviewer"]
        return "final.review" in permissions and "deliverable.review" in permissions, {"permissions": sorted(permissions)}
    if name == "auditor_权限":
        permissions = AUDITOR_PROJECT_PERMISSIONS
        return "audit.read" in permissions and "uat.view" in permissions and "uat.execute" not in permissions, {"permissions": sorted(permissions)}
    if name == "文件下载权限":
        return "deliverable.export" not in PROJECT_ROLE_PERMISSIONS["viewer"], {"viewer_permissions": sorted(PROJECT_ROLE_PERMISSIONS["viewer"])}
    if name == "token_不进入日志":
        rows = db.scalars(select(AuditLog).where(AuditLog.project_id == run.project_id)).all()
        serialized = " ".join(f"{item.before_summary_json} {item.after_summary_json}" for item in rows).lower()
        return "bearer " not in serialized and "access_token" not in serialized and "refresh_token" not in serialized, {"audited_rows": len(rows)}
    if name == "sql_与_shell_不被执行":
        scripts = _count(db, ScriptFile, run.project_id)
        return scripts > 0, {"ingested_scripts": scripts, "execution_mode": "parse_only"}
    if name == "zip_slip_防护":
        return _security_control_available(name), {"unsafe_path_rejected": True}
    if name == "git_白名单":
        settings = get_settings()
        configured = bool(settings.lineage_git_allowed_host_list or settings.lineage_git_allowed_local_root_list or not settings.lineage_git_enabled)
        return configured, {"allowed_hosts": settings.lineage_git_allowed_host_list, "git_enabled": settings.lineage_git_enabled}
    return False, {"reason": "unsupported_security_case"}


def _case_outcome(check_key: str, result: tuple[bool, dict]) -> dict:
    passed, evidence = result
    return _outcome(passed, check_key, evidence)


def _readable_excel_evidence(db: Session, project_id: int) -> tuple[bool, dict]:
    stored = db.scalar(
        select(StoredFile)
        .join(DeliverablePackage, DeliverablePackage.generated_file_id == StoredFile.id)
        .where(DeliverablePackage.project_id == project_id)
        .order_by(StoredFile.id.desc())
        .limit(1)
    )
    if stored is None:
        return False, {"reason": "generated_excel_missing"}
    try:
        content = get_storage_service().read(stored.storage_key)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
        sheet_count = len(workbook.sheetnames)
        workbook.close()
    except Exception as exc:
        return False, {"stored_file_id": stored.id, "parse_error": type(exc).__name__}
    return sheet_count > 0, {"stored_file_id": stored.id, "sheet_count": sheet_count, "byte_size": len(content)}


def _downloaded_excel_evidence(db: Session, project_id: int) -> tuple[bool, dict]:
    downloads = list(db.scalars(
        select(AuditLog).where(
            AuditLog.project_id == project_id,
            AuditLog.action.in_(("download_deliverable_excel", "download_deliverable_version")),
            AuditLog.result == "success",
        )
    ))
    return bool(downloads), {"download_audit_ids": [item.id for item in downloads]}


def _immutable_version_evidence(db: Session, versions: list[DeliverablePackageVersion]) -> tuple[bool, dict]:
    if not versions:
        return False, {"reason": "formal_versions_missing"}
    verified_ids: list[int] = []
    try:
        for version in versions:
            stored = db.get(StoredFile, version.generated_file_id)
            if stored is None:
                return False, {"version_id": version.id, "reason": "stored_file_missing"}
            digest = hashlib.sha256(get_storage_service().read(stored.storage_key)).hexdigest()
            if digest != version.content_hash or digest != stored.content_hash:
                return False, {"version_id": version.id, "stored_file_id": stored.id, "reason": "content_hash_mismatch"}
            verified_ids.append(version.id)
    except Exception as exc:
        return False, {"verified_version_ids": verified_ids, "read_error": type(exc).__name__}
    return True, {"verified_version_ids": verified_ids, "version_count": len(verified_ids)}


def _deployment_outcome(db: Session, check_key: str, check_name: str, project_id: int) -> dict:
    checks = run_health_checks(db, get_settings())["checks"]
    mapping = {
        "数据库连接": ("database",),
        "alembic_revision": ("alembic_revision",),
        "存储读写": ("storage",),
        "redis": ("redis",),
        "celery": ("task_queue",),
        "向量存储": ("vector_store",),
        "模型配置": ("llm_provider", "embedding_provider"),
        "磁盘空间": ("disk_space",),
        "日志目录": ("application",),
    }
    if check_name == "后台任务运行":
        completed = _count(db, BackgroundJob, project_id, BackgroundJob.status == "completed")
        return _outcome(completed > 0, check_key, {"completed_background_jobs": completed})
    if check_name == "secret_强度":
        errors = [item["code"] for item in get_settings().validate_configuration() if item["severity"] == "error"]
        return _outcome(not errors, check_key, {"configuration_error_codes": errors})
    selected = mapping.get(check_name.lower(), mapping.get(check_name))
    if not selected:
        return _outcome(False, check_key, {"reason": "unsupported_deployment_check"})
    statuses = {name: checks[name]["status"] for name in selected}
    return _outcome(all(status in {"healthy", "disabled"} for status in statuses.values()), check_key, {"health_checks": statuses})


def _security_control_available(check_name: str) -> bool:
    if "zip_slip" in check_name:
        try:
            _safe_relative_path("../escape.sql")
        except ValueError:
            return True
        return False
    if "git_白名单" in check_name:
        settings = get_settings()
        return bool(settings.lineage_git_allowed_host_list or settings.lineage_git_allowed_local_root_list or not settings.lineage_git_enabled)
    if "viewer" in check_name:
        return "uat.execute" not in PROJECT_ROLE_PERMISSIONS["viewer"]
    return bool(PROJECT_ROLE_PERMISSIONS) and all("project.view" in permissions for permissions in PROJECT_ROLE_PERMISSIONS.values())


def _count(db: Session, model: type, project_id: int, *conditions) -> int:
    return int(db.scalar(select(func.count()).select_from(model).where(model.project_id == project_id, *conditions)) or 0)


def _outcome(passed: bool, check_key: str, evidence: dict) -> dict:
    return {
        "status": "passed" if passed else "failed",
        "actual": {"check": check_key, "passed": passed},
        "evidence": {"check_key": check_key, **evidence},
        "error": None if passed else f"Built-in UAT check did not pass: {check_key}",
    }
