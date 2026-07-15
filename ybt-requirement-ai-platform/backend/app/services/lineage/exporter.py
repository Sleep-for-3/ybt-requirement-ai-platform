from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    ImpactAnalysis, LineageEdge, LineageNode, ReviewDecision, ReviewTask, ScriptChangeItem,
    ScriptChangeSet, ScriptDependency, ScriptFile, ScriptFileVersion, WorkflowInstance,
)


SHEETS = ["血缘总览", "字段级血缘", "表级血缘", "脚本清单", "脚本依赖", "加工逻辑", "未解析节点", "版本变更", "影响分析", "待确认问题", "审核记录"]


def export_lineage_workbook(db: Session, project_id: int, *, script_file_id: int | None = None, target_field_id: int | None = None) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)
    nodes_query = select(LineageNode).where(LineageNode.project_id == project_id)
    if script_file_id is not None:
        nodes_query = nodes_query.where(LineageNode.script_file_id == script_file_id)
    if target_field_id is not None:
        nodes_query = nodes_query.where(LineageNode.target_field_id == target_field_id)
    nodes = list(db.scalars(nodes_query.order_by(LineageNode.id)).all())
    node_map = {item.id: item for item in nodes}
    edge_query = select(LineageEdge).where(LineageEdge.project_id == project_id, LineageEdge.enabled.is_(True))
    if script_file_id is not None:
        versions = select(ScriptFileVersion.id).where(ScriptFileVersion.script_file_id == script_file_id)
        edge_query = edge_query.where(LineageEdge.script_file_version_id.in_(versions))
    if target_field_id is not None and node_map:
        ids = list(node_map)
        edge_query = edge_query.where((LineageEdge.source_node_id.in_(ids)) | (LineageEdge.target_node_id.in_(ids)))
    edges = list(db.scalars(edge_query.order_by(LineageEdge.id)).all())
    for edge in edges:
        if edge.source_node_id not in node_map:
            row = db.get(LineageNode, edge.source_node_id)
            if row: node_map[row.id] = row
        if edge.target_node_id not in node_map:
            row = db.get(LineageNode, edge.target_node_id)
            if row: node_map[row.id] = row

    _write(workbook["血缘总览"], ["指标", "数量"], [
        ["节点", len(nodes)], ["边", len(edges)], ["未解析节点", sum(1 for item in nodes if item.unresolved_flag)],
        ["脚本", db.query(ScriptFile).filter(ScriptFile.project_id == project_id).count()],
        ["影响分析", db.query(ImpactAnalysis).filter(ImpactAnalysis.project_id == project_id).count()],
    ])
    field_headers = ["一表通表", "一表通字段", "监管集市表", "监管集市字段", "贴源层系统", "来源库", "来源 schema", "来源表", "来源字段", "处理逻辑", "过滤条件", "关联条件", "码值转换", "脚本文件", "脚本版本", "源码行号", "置信度", "血缘状态", "最近验证时间"]
    field_rows = []
    table_rows = []
    logic_rows = []
    for edge in edges:
        source = node_map.get(edge.source_node_id); target = node_map.get(edge.target_node_id)
        version = db.get(ScriptFileVersion, edge.script_file_version_id)
        script = db.get(ScriptFile, version.script_file_id) if version else None
        if source and target:
            if source.column_name or target.column_name:
                field_rows.append([
                    target.table_name if target.target_field_id else "", target.column_name if target.target_field_id else "",
                    target.table_name if target.mart_field_id else "", target.column_name if target.mart_field_id else "",
                    "", source.database_name or "", source.schema_name or "", source.table_name or "", source.column_name or "",
                    edge.transformation_expression or "", edge.filter_condition or "", edge.join_condition or "", edge.code_mapping_rule or "",
                    script.relative_path if script else "", version.version_no if version else "", _line_range(edge), edge.confidence_level,
                    "verified" if not target.unresolved_flag else "not_linked", "",
                ])
            if source.table_name and target.table_name:
                table_rows.append([source.logical_name, target.logical_name, edge.edge_type, script.relative_path if script else "", version.version_no if version else "", edge.confidence_level])
        logic_rows.append([edge.id, source.logical_name if source else "", target.logical_name if target else "", edge.edge_type, edge.transformation_expression or "", edge.filter_condition or "", edge.join_condition or "", edge.aggregation_rule or "", edge.code_mapping_rule or "", _line_range(edge)])
    _write(workbook["字段级血缘"], field_headers, field_rows)
    _write(workbook["表级血缘"], ["来源表", "目标表", "关系类型", "脚本文件", "版本", "置信度"], table_rows)

    scripts_query = select(ScriptFile).where(ScriptFile.project_id == project_id)
    if script_file_id is not None: scripts_query = scripts_query.where(ScriptFile.id == script_file_id)
    scripts = list(db.scalars(scripts_query.order_by(ScriptFile.relative_path)).all())
    _write(workbook["脚本清单"], ["脚本 ID", "相对路径", "类型", "当前版本", "状态", "逻辑目标"], [[item.id, item.relative_path, item.file_type, item.current_version_no, "enabled" if item.enabled else "deleted", item.logical_target_name or ""] for item in scripts])
    script_ids = [item.id for item in scripts]
    dependencies = list(db.scalars(select(ScriptDependency).where(ScriptDependency.parent_script_file_id.in_(script_ids)).order_by(ScriptDependency.id)).all()) if script_ids else []
    _write(workbook["脚本依赖"], ["父脚本 ID", "子脚本 ID", "依赖类型", "调用表达式", "条件", "源码行", "置信度", "warning"], [[item.parent_script_file_id, item.child_script_file_id or "", item.dependency_type, item.call_expression, item.condition_expression or "", item.source_line_start or "", item.confidence_level, "；".join(item.warnings_json)] for item in dependencies])
    _write(workbook["加工逻辑"], ["边 ID", "来源", "目标", "类型", "转换表达式", "过滤条件", "关联条件", "聚合", "码值映射", "源码行"], logic_rows)
    _write(workbook["未解析节点"], ["节点 ID", "类型", "逻辑名称", "库", "schema", "表", "字段", "脚本版本"], [[item.id, item.node_type, item.logical_name, item.database_name or "", item.schema_name or "", item.table_name or "", item.column_name or "", item.script_file_version_id or ""] for item in nodes if item.unresolved_flag])

    changes_query = select(ScriptChangeSet).where(ScriptChangeSet.project_id == project_id)
    if script_file_id is not None: changes_query = changes_query.where(ScriptChangeSet.script_file_id == script_file_id)
    changes = list(db.scalars(changes_query.order_by(ScriptChangeSet.id.desc())).all())
    change_rows = []
    for change in changes:
        for item in db.scalars(select(ScriptChangeItem).where(ScriptChangeItem.change_set_id == change.id).order_by(ScriptChangeItem.id)).all():
            change_rows.append([change.id, change.script_file_id, change.from_version_id or "", change.to_version_id or "", item.change_category, item.entity_type, item.severity, str(item.old_value_json), str(item.new_value_json)])
    _write(workbook["版本变更"], ["变更集", "脚本", "旧版本", "新版本", "类别", "实体", "严重度", "旧值", "新值"], change_rows)
    change_ids = [item.id for item in changes]
    impacts = list(db.scalars(select(ImpactAnalysis).where(ImpactAnalysis.change_set_id.in_(change_ids)).order_by(ImpactAnalysis.id.desc())).all()) if change_ids else []
    _write(workbook["影响分析"], ["影响 ID", "变更集", "状态", "严重度", "一表通字段", "集市字段", "受影响口径", "摘要"], [[item.id, item.change_set_id, item.status, item.severity, str(item.affected_target_field_ids_json), str(item.affected_mart_field_ids_json), str(item.affected_mapping_ids_json), str(item.summary_json)] for item in impacts])
    _write(workbook["待确认问题"], ["影响 ID", "严重度", "问题"], [[item.id, item.severity, question] for item in impacts for question in item.open_questions_json])
    instances = list(db.scalars(select(WorkflowInstance).where(WorkflowInstance.project_id == project_id, WorkflowInstance.workflow_key == "lineage_change_review")).all())
    instance_ids = [item.id for item in instances]
    tasks = list(db.scalars(select(ReviewTask).where(ReviewTask.workflow_instance_id.in_(instance_ids))).all()) if instance_ids else []
    task_ids = [item.id for item in tasks]
    decisions = list(db.scalars(select(ReviewDecision).where(ReviewDecision.review_task_id.in_(task_ids)).order_by(ReviewDecision.id)).all()) if task_ids else []
    task_map = {item.id: item for item in tasks}
    _write(workbook["审核记录"], ["任务 ID", "步骤", "决定", "意见", "审核人", "审核时间"], [[item.review_task_id, task_map[item.review_task_id].step_key, item.decision, item.comment or "", item.decided_by, item.decided_at] for item in decisions])
    stream = BytesIO(); workbook.save(stream); return stream.getvalue()


def _write(sheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows: sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    for row in sheet.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        values = [str(sheet.cell(row=row, column=index).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values), default=len(header)) + 2, 12), 45)


def _line_range(edge: LineageEdge) -> str:
    if edge.source_line_start is None: return ""
    return str(edge.source_line_start) if edge.source_line_end in {None, edge.source_line_start} else f"{edge.source_line_start}-{edge.source_line_end}"
