from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter


def inspect_workbook(content: bytes) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        formulas = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.data_type == "f"]
        used_headers = _detect_header_rows(sheet)
        sheets.append({
            "sheet_name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "used_range": f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
            "header_row_start": used_headers[0],
            "header_row_end": used_headers[1],
            "merged_cells": [str(item) for item in sheet.merged_cells.ranges],
            "hidden_rows": [index for index, dimension in sheet.row_dimensions.items() if dimension.hidden],
            "hidden_columns": [index for index, dimension in sheet.column_dimensions.items() if dimension.hidden],
            "freeze_panes": str(sheet.freeze_panes or ""),
            "data_validation_count": len(sheet.data_validations.dataValidation),
            "formula_cells": formulas[:200],
            "styled_cell_count": sum(1 for row in sheet.iter_rows() for cell in row if cell.has_style),
            "row_heights": {str(i): d.height for i, d in sheet.row_dimensions.items() if d.height},
            "column_widths": {i: d.width for i, d in sheet.column_dimensions.items() if d.width},
            "sample_rows": [[sheet.cell(row, col).value for col in range(1, min(sheet.max_column, 20) + 1)] for row in range(1, min(sheet.max_row, 5) + 1)],
        })
    return {"sheet_names": workbook.sheetnames, "sheets": sheets}


def render_workbook(content: bytes, sheet_mappings: list[Any], column_mappings: list[Any], records: dict[str, list[dict]], *, preview_limit: int | None = None) -> tuple[bytes, list[dict]]:
    """Render business records through one narrow interface while preserving workbook layout."""
    workbook = load_workbook(BytesIO(content), data_only=False)
    warnings: list[dict] = []
    columns_by_sheet: dict[int, list[Any]] = {}
    for column in column_mappings:
        columns_by_sheet.setdefault(column.template_sheet_mapping_id, []).append(column)
    for mapping in sheet_mappings:
        if not mapping.enabled:
            continue
        if mapping.sheet_name not in workbook.sheetnames:
            warnings.append({"type": "missing_sheet", "sheet": mapping.sheet_name})
            continue
        sheet = workbook[mapping.sheet_name]
        rows = list(records.get(mapping.business_section, []))
        if preview_limit is not None:
            rows = rows[:preview_limit]
        mapped_columns = columns_by_sheet.get(mapping.id, [])
        for offset, record in enumerate(rows):
            row_number = mapping.data_start_row + offset
            _copy_template_row(sheet, mapping.data_start_row, row_number)
            for column in mapped_columns:
                col_number = column_index_from_string(column.excel_column)
                cell = sheet.cell(row_number, col_number)
                if isinstance(cell, MergedCell):
                    warnings.append({"type": "merge_conflict", "sheet": sheet.title, "cell": f"{column.excel_column}{row_number}"})
                    continue
                value = record.get(column.business_field, column.default_value)
                if column.required and value in (None, ""):
                    warnings.append({"type": "required_missing", "sheet": sheet.title, "cell": cell.coordinate, "business_field": column.business_field})
                if cell.data_type == "f" and column.write_mode != "fill_blank_only":
                    warnings.append({"type": "formula_overlap", "sheet": sheet.title, "cell": cell.coordinate})
                    continue
                if column.write_mode == "fill_blank_only" and cell.value not in (None, ""):
                    continue
                cell.value = value
                if isinstance(value, str) and len(value) > 30:
                    alignment = copy(cell.alignment)
                    alignment.wrap_text = True
                    alignment.vertical = "top"
                    cell.alignment = alignment
                    sheet.row_dimensions[row_number].height = max(sheet.row_dimensions[row_number].height or 15, min(120, 15 + len(value) // 20 * 12))
            _apply_merges(sheet, mapping, mapped_columns, row_number, record, rows, offset, warnings)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), warnings


def _detect_header_rows(sheet) -> tuple[int, int]:
    scored = []
    for row in range(1, min(sheet.max_row, 15) + 1):
        values = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        text_count = sum(bool(isinstance(value, str) and value.strip()) for value in values)
        scored.append((text_count, row))
    best = max(scored, default=(0, 1))[1]
    merged_end = max((item.max_row for item in sheet.merged_cells.ranges if item.min_row <= best <= item.max_row), default=best)
    return best, merged_end


def _copy_template_row(sheet, source_row: int, target_row: int) -> None:
    if source_row == target_row or source_row > sheet.max_row:
        return
    source_dimension = sheet.row_dimensions[source_row]
    if source_dimension.height:
        sheet.row_dimensions[target_row].height = source_dimension.height
    for col in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def _apply_merges(sheet, mapping, columns, row_number, record, rows, offset, warnings) -> None:
    if offset == 0:
        return
    previous = rows[offset - 1]
    for column in columns:
        key = "target_field_code" if column.merge_strategy == "merge_same_target_field" else "scenario_code" if column.merge_strategy == "merge_same_scenario" else None
        if key and previous.get(key) == record.get(key) and record.get(key) is not None:
            col = column_index_from_string(column.excel_column)
            try:
                sheet.merge_cells(start_row=row_number - 1, start_column=col, end_row=row_number, end_column=col)
            except ValueError:
                warnings.append({"type": "merge_conflict", "sheet": sheet.title, "cell": f"{column.excel_column}{row_number}"})
