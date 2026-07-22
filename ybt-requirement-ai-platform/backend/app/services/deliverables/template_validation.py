from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy import select

from app.models import (
    DeliverableTemplate,
    DeliverableTemplateVersion,
    StoredFile,
    TemplateColumnMapping,
    TemplateSheetMapping,
)
from app.services.storage import get_storage_service


REQUIRED_SECTIONS_BY_TEMPLATE_TYPE: dict[str, set[str]] = {
    "business_traceability": {"target_field", "scenario_business_mapping", "scenario_technical_lineage"},
    "source_to_mart": {"source_to_mart"},
    "mart_to_ybt": {"mart_to_ybt"},
    "full_delivery_package": {
        "target_field",
        "scenario_business_mapping",
        "scenario_technical_lineage",
        "source_to_mart",
        "mart_to_ybt",
        "evidence",
        "pending_question",
        "review_record",
        "lineage",
        "change_impact",
    },
    "pending_questions": {"pending_question"},
    "evidence_matrix": {"evidence"},
    "change_comparison": {"change_impact"},
}

BUSINESS_FIELDS_BY_SECTION: dict[str, set[str]] = {
    "target_field": {
        "target_table_code", "target_table_name", "target_field_code", "target_field_name",
        "regulatory_definition", "data_type", "field_order",
    },
    "scenario_business_mapping": {
        "target_field_code", "target_field_name", "scenario_code", "scenario_name",
        "business_final_content", "business_ai_draft", "business_confirm_status",
        "business_confidence_level", "business_open_questions",
    },
    "scenario_technical_lineage": {
        "target_field_code", "target_field_name", "scenario_code", "scenario_name",
        "technical_final_content", "source_system_name", "database_name", "schema_name",
        "source_table_name", "source_field_name", "technical_confirm_status", "lineage_status",
    },
    "source_to_mart": {
        "mapping_id", "source_to_mart_final_content", "source_to_mart_status", "source_system_name",
        "source_field_name", "filter_condition", "join_condition", "code_mapping_rule",
        "priority_rule", "null_handling_rule",
    },
    "mart_to_ybt": {
        "mapping_id", "target_field_id", "mart_to_ybt_final_content", "mart_to_ybt_status",
        "filter_condition", "join_condition", "code_mapping_rule", "null_handling_rule",
    },
    "pending_question": {
        "id", "question_type", "question_text", "question_status", "priority", "assigned_role",
        "resolution_text", "target_table_id", "target_field_id", "scenario_id", "source_type", "source_id",
    },
    "evidence": {
        "target_field_id", "evidence_type", "evidence_source", "evidence_location",
        "evidence_summary", "citation", "claim_type",
    },
    "review_record": {
        "action", "resource_type", "resource_id", "reviewer", "approved_at", "review_comment",
    },
    "lineage": {
        "source_script", "script_version", "source_database", "source_schema", "source_table",
        "source_column", "target_database", "target_schema", "target_table", "target_column",
        "edge_type", "transformation_summary", "filter_summary", "join_summary", "lineage_status",
        "reviewed_status", "reviewed_at", "affected_target_field_code", "affected_mapping_type",
        "affected_mapping_id",
    },
    "change_impact": {
        "script_path", "old_version_no", "new_version_no", "change_type", "impact_severity",
        "impact_status", "affected_target_table", "affected_target_field", "affected_scenario",
        "affected_source_to_mart_mapping", "affected_mart_to_ybt_mapping", "change_summary",
        "review_decision", "reviewer", "reviewed_at",
    },
}


def validate_template_version(db, version_id: int) -> dict:
    version = db.get(DeliverableTemplateVersion, version_id)
    if version is None:
        return _result([_issue("error", "template_version_missing", "模板版本不存在")])
    template = db.get(DeliverableTemplate, version.template_id)
    if template is None or template.project_id != version.project_id:
        return _result([_issue("error", "template_missing", "模板不存在或不属于当前项目")])
    stored = db.get(StoredFile, version.stored_file_id)
    issues: list[dict] = []
    workbook = None
    if stored is None or stored.project_id != version.project_id or not stored.enabled:
        issues.append(_issue("error", "template_file_missing", "模板原文件不存在或已停用"))
    else:
        try:
            workbook = load_workbook(BytesIO(get_storage_service().read(stored.storage_key)), data_only=False)
        except Exception:
            issues.append(_issue("error", "template_file_invalid", "模板原文件无法读取"))

    mappings = list(db.scalars(select(TemplateSheetMapping).where(
        TemplateSheetMapping.project_id == version.project_id,
        TemplateSheetMapping.template_version_id == version.id,
    ).order_by(TemplateSheetMapping.id)).all())
    enabled = [mapping for mapping in mappings if mapping.enabled]
    if not enabled:
        issues.append(_issue("error", "sheet_mapping_missing", "至少启用一个 Sheet 映射"))
    configured_sections = {mapping.business_section for mapping in enabled}
    for section in sorted(REQUIRED_SECTIONS_BY_TEMPLATE_TYPE.get(template.template_type, set()) - configured_sections):
        issues.append(_issue("error", "required_section_missing", f"缺少必需业务区域：{section}", business_section=section))

    seen_sections: set[tuple[str, str]] = set()
    for mapping in enabled:
        scope = (mapping.business_section, mapping.sheet_name)
        if scope in seen_sections:
            issues.append(_issue("error", "duplicate_sheet_section", "同一业务区域和 Sheet 不能重复配置", mapping))
        seen_sections.add(scope)
        if workbook is not None and mapping.sheet_name not in workbook.sheetnames:
            issues.append(_issue("error", "missing_sheet", "配置的 Sheet 不存在", mapping))
            continue
        if mapping.header_row_start < 1 or mapping.header_row_end < mapping.header_row_start:
            issues.append(_issue("error", "invalid_header_rows", "模板表头行配置无效", mapping))
        if mapping.data_start_row <= mapping.header_row_end:
            issues.append(_issue("error", "invalid_data_start_row", "数据起始行不得覆盖模板表头", mapping))
        columns = list(db.scalars(select(TemplateColumnMapping).where(
            TemplateColumnMapping.project_id == version.project_id,
            TemplateColumnMapping.template_sheet_mapping_id == mapping.id,
        ).order_by(TemplateColumnMapping.id)).all())
        if not columns:
            issues.append(_issue("error", "column_mapping_missing", "业务区域至少需要一个列映射", mapping))
            continue
        seen_columns: set[str] = set()
        seen_fields: set[str] = set()
        for column in columns:
            if column.excel_column in seen_columns or column.business_field in seen_fields:
                issues.append(_issue("error", "duplicate_column_mapping", "同一业务区域存在重复列或字段映射", mapping, column))
            seen_columns.add(column.excel_column)
            seen_fields.add(column.business_field)
            if column.business_field not in BUSINESS_FIELDS_BY_SECTION.get(mapping.business_section, set()):
                issues.append(_issue("error", "invalid_business_field", "business_field 不在允许集合中", mapping, column))
            try:
                column_number = column_index_from_string(column.excel_column)
            except ValueError:
                column_number = 0
            if not 1 <= column_number <= 16384:
                issues.append(_issue("error", "invalid_column_mapping", "Excel 列必须位于 A 到 XFD", mapping, column))
                continue
            if workbook is not None and mapping.sheet_name in workbook.sheetnames:
                cell = workbook[mapping.sheet_name].cell(mapping.data_start_row, column_number)
                if column.required and cell.data_type == "f" and column.write_mode != "fill_blank_only":
                    issues.append(_issue(
                        "error",
                        "required_field_formula_overlap",
                        "必填业务字段不能覆盖模板公式单元格",
                        mapping,
                        column,
                        cell.coordinate,
                    ))
    return _result(issues)


def _result(issues: list[dict]) -> dict:
    return {
        "valid": not any(issue["severity"] == "error" for issue in issues),
        "error_count": sum(issue["severity"] == "error" for issue in issues),
        "warning_count": sum(issue["severity"] == "warning" for issue in issues),
        "issues": issues,
    }


def _issue(
    severity: str,
    code: str,
    message: str,
    mapping=None,
    column=None,
    cell: str | None = None,
    *,
    business_section: str | None = None,
) -> dict:
    return {
        "severity": severity,
        "code": code,
        "sheet_name": getattr(mapping, "sheet_name", None),
        "cell": cell,
        "business_section": business_section or getattr(mapping, "business_section", None),
        "business_field": getattr(column, "business_field", None),
        "message": message,
    }
