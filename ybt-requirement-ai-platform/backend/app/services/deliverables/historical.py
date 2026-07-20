import hashlib
import re
import unicodedata
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from app.models import HistoricalCaliberItem, ProductScenario, TargetField, TargetTable


ALIASES = {
    "target_table_code": {"表代码", "表英文名", "target_table_code"},
    "target_field_code": {"字段代码", "字段英文名", "数据项编码", "target_field_code"},
    "target_field_name": {"字段名称", "字段中文名", "数据项名称", "target_field_name"},
    "scenario_name": {"场景", "业务场景", "产品场景", "scenario_name"},
    "business_content": {"业务口径", "业务定义", "business_content"},
    "technical_content": {"技术溯源", "处理逻辑", "technical_content"},
    "source_system_name": {"来源系统", "源系统"}, "database_name": {"来源库", "数据库"},
    "schema_name": {"schema", "来源schema"}, "source_table_name": {"来源表", "来源表英文名"},
    "source_field_name": {"来源字段", "来源字段英文名"}, "mart_table_name": {"集市表"},
    "mart_field_name": {"集市字段"}, "filter_condition": {"过滤条件"}, "join_condition": {"join条件", "关联条件"},
    "code_mapping_rule": {"码值映射", "码值转换"}, "priority_rule": {"优先级规则"}, "null_handling_rule": {"空值处理"},
}


def parse_historical_workbook(content: bytes, db, project_id: int, import_id: int) -> tuple[list[HistoricalCaliberItem], list[dict]]:
    workbook = load_workbook(BytesIO(content), data_only=False)
    items, warnings = [], []
    for sheet in workbook.worksheets:
        header_row, mapping = _find_header(sheet)
        if not mapping:
            warnings.append({"sheet": sheet.title, "message": "未识别到历史口径表头"})
            continue
        for row in range(header_row + 1, sheet.max_row + 1):
            values = {field: sheet.cell(row, col).value for field, col in mapping.items()}
            if not any(value not in (None, "") for value in values.values()):
                continue
            field_matches = list(db.scalars(select(TargetField).where(TargetField.project_id == project_id, TargetField.field_code == str(values.get("target_field_code") or "").strip())).all())
            if not field_matches and values.get("target_field_name"):
                field_matches = list(db.scalars(select(TargetField).where(TargetField.project_id == project_id, TargetField.field_name == str(values["target_field_name"]).strip())).all())
            scenarios = list(db.scalars(select(ProductScenario).where(ProductScenario.project_id == project_id, ProductScenario.scenario_name == str(values.get("scenario_name") or "").strip())).all()) if values.get("scenario_name") else []
            match_status = "matched" if len(field_matches) == 1 and len(scenarios) <= 1 else "ambiguous" if len(field_matches) > 1 or len(scenarios) > 1 else "unmatched"
            payload = "|".join(str(values.get(field) or "") for field in sorted(ALIASES))
            item = HistoricalCaliberItem(project_id=project_id, historical_import_id=import_id, source_sheet_name=sheet.title, source_cell_range=f"A{row}:{sheet.cell(row, sheet.max_column).coordinate}", content_hash=hashlib.sha256(payload.encode()).hexdigest(), match_status=match_status, matched_target_field_id=field_matches[0].id if len(field_matches) == 1 else None, matched_scenario_id=scenarios[0].id if len(scenarios) == 1 else None, **values)
            items.append(item)
    return items, warnings


def semantic_diff(left: str | None, right: str | None) -> str:
    a, b = _normalize(left), _normalize(right)
    if a == b:
        return "unchanged"
    if not a:
        return "added"
    if not b:
        return "removed"
    return "modified"


def _normalize(value: str | None) -> str:
    text = unicodedata.normalize("NFKC", value or "").lower()
    return re.sub(r"[\s，。；：、,.!！?？:;()（）]+", "", text)


def _find_header(sheet) -> tuple[int, dict[str, int]]:
    for row in range(1, min(sheet.max_row, 12) + 1):
        mapping = {}
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(row, col).value or "").strip().lower()
            for field, aliases in ALIASES.items():
                if header in {alias.lower() for alias in aliases}:
                    mapping[field] = col
        if mapping.get("target_field_code") or mapping.get("target_field_name"):
            return row, mapping
    return 1, {}
